"""
Comprehensive Test Suite for Reporting Module

Tests cover:
1. Schema inference and column classification
2. Aggregation logic
3. Template analysis (placeholder vs column mode)
4. Excel report generation
5. CSV report generation
6. Cache functionality
7. Integration with streamlit_reporting.py
"""

import io
import os
import sys
import tempfile
import shutil
from pathlib import Path
from datetime import date, datetime
from unittest.mock import MagicMock, patch

import pytest
import pandas as pd
import numpy as np

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


# =============================================================================
# FIXTURES
# =============================================================================

@pytest.fixture
def sample_campaign_df():
    """Create sample campaign DataFrame for testing."""
    return pd.DataFrame({
        'Date': pd.date_range('2024-07-01', periods=20, freq='D'),
        'Platform': ['Google Ads', 'Meta Ads', 'Snapchat Ads', 'YouTube', 'Bing Ads'] * 4,
        'Campaign_Name': [f'Campaign_{i}' for i in range(20)],
        'AdGroup_Name': [f'AdGroup_{i % 5}' for i in range(20)],
        'Ad_Name': [f'Ad_{i}' for i in range(20)],
        'Impressions': np.random.randint(10000, 500000, 20),
        'Clicks': np.random.randint(100, 10000, 20),
        'Spend_USD': np.random.uniform(100, 5000, 20).round(2),
        'Conversions': np.random.randint(10, 500, 20),
        'Revenue_USD': np.random.uniform(500, 25000, 20).round(2),
    })


@pytest.fixture
def sample_pacing_df():
    """Create sample pacing DataFrame for testing."""
    return pd.DataFrame({
        'Campaign_ID': [f'CMP-{i:03d}' for i in range(10)],
        'Campaign_Name': [f'Campaign_{i}' for i in range(10)],
        'Platform': ['Google Ads', 'Meta Ads'] * 5,
        'Date': pd.date_range('2024-10-01', periods=10, freq='D'),
        'Monthly_Budget_USD': [10000] * 10,
        'Daily_Target_Budget_USD': [322.58] * 10,
        'Actual_Daily_Spend_USD': np.random.uniform(280, 360, 10).round(2),
    })


@pytest.fixture
def temp_dir():
    """Create temporary directory for test files."""
    tmpdir = tempfile.mkdtemp()
    yield tmpdir
    shutil.rmtree(tmpdir, ignore_errors=True)


@pytest.fixture
def sample_csv_file(temp_dir, sample_campaign_df):
    """Create sample CSV file."""
    csv_path = Path(temp_dir) / "test_data.csv"
    sample_campaign_df.to_csv(csv_path, index=False)
    return csv_path


@pytest.fixture
def sample_excel_template(temp_dir):
    """Create sample Excel template with column headers."""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Pacing"
    
    # Add headers
    headers = ['Date', 'Platform', 'Impressions', 'Clicks', 'CTR', 
               'Conversions', 'Conv_Rate', 'Spend', 'CPM', 'CPC', 'CPA', 'ROAS', 'Revenue']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add Weekly Pacing sheet
    ws2 = wb.create_sheet("Weekly Pacing")
    weekly_headers = ['Week', 'Week_Start', 'Week_End', 'Platform', 'Impressions', 
                      'Clicks', 'CTR', 'Conversions', 'Conv_Rate', 'Spend', 
                      'CPM', 'CPC', 'CPA', 'ROAS', 'Revenue']
    for col, header in enumerate(weekly_headers, 1):
        ws2.cell(row=1, column=col, value=header)
    
    excel_path = Path(temp_dir) / "test_template.xlsx"
    wb.save(excel_path)
    return excel_path


@pytest.fixture
def placeholder_excel_template(temp_dir):
    """Create Excel template with placeholders."""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = "Campaign Report"
    ws['A3'] = "Total Spend:"
    ws['B3'] = "{{Total_Spend}}"
    ws['A4'] = "Total Revenue:"
    ws['B4'] = "{{Total_Revenue}}"
    ws['A5'] = "Total Clicks:"
    ws['B5'] = "{{Total_Clicks}}"
    
    excel_path = Path(temp_dir) / "placeholder_template.xlsx"
    wb.save(excel_path)
    return excel_path


# =============================================================================
# SCHEMA INFERENCE TESTS
# =============================================================================

class TestSchemaInference:
    """Tests for column type detection and schema inference."""
    
    def test_detect_date_columns(self, sample_campaign_df):
        """Test detection of date columns."""
        date_cols = []
        for col in sample_campaign_df.columns:
            if pd.api.types.is_datetime64_any_dtype(sample_campaign_df[col]):
                date_cols.append(col)
        
        assert 'Date' in date_cols
        assert len(date_cols) == 1
    
    def test_detect_numeric_columns(self, sample_campaign_df):
        """Test detection of numeric columns."""
        numeric_cols = sample_campaign_df.select_dtypes(include=['number']).columns.tolist()
        
        assert 'Impressions' in numeric_cols
        assert 'Clicks' in numeric_cols
        assert 'Spend_USD' in numeric_cols
        assert 'Conversions' in numeric_cols
        assert 'Revenue_USD' in numeric_cols
    
    def test_detect_text_columns(self, sample_campaign_df):
        """Test detection of text/categorical columns."""
        text_cols = sample_campaign_df.select_dtypes(include=['object']).columns.tolist()
        
        assert 'Platform' in text_cols
        assert 'Campaign_Name' in text_cols
        assert 'AdGroup_Name' in text_cols
    
    def test_infer_aggregation_type(self, sample_campaign_df):
        """Test inference of appropriate aggregation methods."""
        # Metrics that should be summed
        sum_cols = ['Impressions', 'Clicks', 'Spend_USD', 'Conversions', 'Revenue_USD']
        
        # Verify these are numeric and can be summed
        for col in sum_cols:
            assert col in sample_campaign_df.columns
            assert pd.api.types.is_numeric_dtype(sample_campaign_df[col])
            assert sample_campaign_df[col].sum() > 0
    
    def test_detect_currency_columns(self, sample_campaign_df):
        """Test detection of currency columns by naming convention."""
        currency_cols = [col for col in sample_campaign_df.columns if 'USD' in col or 'Spend' in col or 'Revenue' in col]
        
        assert 'Spend_USD' in currency_cols
        assert 'Revenue_USD' in currency_cols
    
    def test_detect_rate_columns(self):
        """Test detection of rate/percentage columns."""
        df = pd.DataFrame({
            'CTR_%': [1.5, 2.0, 1.8],
            'Conversion_Rate_%': [5.0, 4.5, 6.0],
            'ROAS': [3.5, 4.0, 3.8],
        })
        
        rate_cols = [col for col in df.columns if '%' in col or 'Rate' in col.lower()]
        assert 'CTR_%' in rate_cols
        assert 'Conversion_Rate_%' in rate_cols


# =============================================================================
# AGGREGATION TESTS
# =============================================================================

class TestAggregation:
    """Tests for data aggregation logic."""
    
    def test_daily_aggregation(self, sample_campaign_df):
        """Test daily aggregation by date and platform."""
        daily_agg = sample_campaign_df.groupby(['Date', 'Platform']).agg({
            'Impressions': 'sum',
            'Clicks': 'sum',
            'Spend_USD': 'sum',
            'Conversions': 'sum',
            'Revenue_USD': 'sum'
        }).reset_index()
        
        assert len(daily_agg) <= len(sample_campaign_df)
        assert 'Date' in daily_agg.columns
        assert 'Platform' in daily_agg.columns
        assert daily_agg['Impressions'].sum() == sample_campaign_df['Impressions'].sum()
    
    def test_weekly_aggregation(self, sample_campaign_df):
        """Test weekly aggregation."""
        df = sample_campaign_df.copy()
        df['Week'] = df['Date'].dt.isocalendar().week
        df['Year'] = df['Date'].dt.isocalendar().year
        
        weekly_agg = df.groupby(['Year', 'Week']).agg({
            'Date': ['min', 'max'],
            'Impressions': 'sum',
            'Clicks': 'sum',
            'Spend_USD': 'sum',
        }).reset_index()
        
        assert len(weekly_agg) <= len(sample_campaign_df)
        assert weekly_agg[('Impressions', 'sum')].sum() == sample_campaign_df['Impressions'].sum()
    
    def test_platform_aggregation(self, sample_campaign_df):
        """Test aggregation by platform."""
        platform_agg = sample_campaign_df.groupby('Platform').agg({
            'Impressions': 'sum',
            'Clicks': 'sum',
            'Spend_USD': 'sum',
            'Revenue_USD': 'sum'
        }).reset_index()
        
        assert len(platform_agg) == sample_campaign_df['Platform'].nunique()
        assert platform_agg['Spend_USD'].sum() == pytest.approx(sample_campaign_df['Spend_USD'].sum(), rel=1e-5)
    
    def test_calculated_metrics(self, sample_campaign_df):
        """Test calculation of derived metrics."""
        df = sample_campaign_df.copy()
        
        # CTR
        df['CTR'] = (df['Clicks'] / df['Impressions']) * 100
        assert df['CTR'].notna().all()
        assert (df['CTR'] >= 0).all()
        
        # CPC
        df['CPC'] = df['Spend_USD'] / df['Clicks']
        assert df['CPC'].notna().all()
        
        # ROAS
        df['ROAS'] = df['Revenue_USD'] / df['Spend_USD']
        assert df['ROAS'].notna().all()
    
    def test_handle_zero_division(self):
        """Test handling of division by zero in calculations."""
        df = pd.DataFrame({
            'Clicks': [100, 0, 50],
            'Impressions': [1000, 0, 500],
            'Spend': [100, 0, 50],
        })
        
        # Safe CTR calculation
        df['CTR'] = df.apply(lambda r: 0 if r['Impressions'] == 0 else r['Clicks'] / r['Impressions'], axis=1)
        assert df['CTR'].iloc[1] == 0
        
        # Safe CPC calculation
        df['CPC'] = df.apply(lambda r: 0 if r['Clicks'] == 0 else r['Spend'] / r['Clicks'], axis=1)
        assert df['CPC'].iloc[1] == 0


# =============================================================================
# TEMPLATE ANALYSIS TESTS
# =============================================================================

class TestTemplateAnalysis:
    """Tests for template structure analysis."""
    
    def test_analyze_column_mode_template(self, sample_excel_template):
        """Test analysis of column-based Excel template."""
        from openpyxl import load_workbook
        
        wb = load_workbook(sample_excel_template)
        sheet = wb.active
        
        # Extract headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append({
                    'name': str(cell.value),
                    'column_letter': cell.column_letter,
                    'sheet': sheet.title
                })
        
        assert len(headers) > 0
        assert any(h['name'] == 'Date' for h in headers)
        assert any(h['name'] == 'Platform' for h in headers)
        assert any(h['name'] == 'Impressions' for h in headers)
    
    def test_analyze_placeholder_template(self, placeholder_excel_template):
        """Test analysis of placeholder-based Excel template."""
        from openpyxl import load_workbook
        
        wb = load_workbook(placeholder_excel_template)
        sheet = wb.active
        
        placeholders = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if '{{' in cell.value:
                        placeholders.append({
                            'name': cell.value,
                            'location': f"{sheet.title}!{cell.coordinate}"
                        })
        
        assert len(placeholders) == 3
        assert any('Total_Spend' in p['name'] for p in placeholders)
        assert any('Total_Revenue' in p['name'] for p in placeholders)
    
    def test_detect_template_mode(self, sample_excel_template, placeholder_excel_template):
        """Test automatic detection of template mode."""
        from openpyxl import load_workbook
        
        # Column mode template
        wb1 = load_workbook(sample_excel_template)
        has_placeholders = False
        for sheet in wb1.sheetnames:
            for row in wb1[sheet].iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                        has_placeholders = True
        
        assert not has_placeholders  # Should be column mode
        
        # Placeholder mode template
        wb2 = load_workbook(placeholder_excel_template)
        has_placeholders = False
        for sheet in wb2.sheetnames:
            for row in wb2[sheet].iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                        has_placeholders = True
        
        assert has_placeholders  # Should be placeholder mode
    
    def test_multi_sheet_template(self, sample_excel_template):
        """Test analysis of multi-sheet template."""
        from openpyxl import load_workbook
        
        wb = load_workbook(sample_excel_template)
        
        assert len(wb.sheetnames) >= 2
        assert 'Daily Pacing' in wb.sheetnames
        assert 'Weekly Pacing' in wb.sheetnames


# =============================================================================
# EXCEL REPORT GENERATION TESTS
# =============================================================================

class TestExcelReportGeneration:
    """Tests for Excel report generation."""
    
    def test_generate_column_mode_report(self, sample_excel_template, sample_campaign_df, temp_dir):
        """Test generation of column-mode Excel report."""
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = load_workbook(sample_excel_template)
        sheet = wb['Daily Pacing']
        
        # Write data starting from row 2
        for r_idx, row in enumerate(dataframe_to_rows(sample_campaign_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        output_path = Path(temp_dir) / "output_report.xlsx"
        wb.save(output_path)
        
        assert output_path.exists()
        
        # Verify data was written
        wb_check = load_workbook(output_path)
        sheet_check = wb_check['Daily Pacing']
        assert sheet_check.cell(row=2, column=1).value is not None
    
    def test_generate_placeholder_report(self, placeholder_excel_template, sample_campaign_df, temp_dir):
        """Test generation of placeholder-mode Excel report."""
        from openpyxl import load_workbook
        
        wb = load_workbook(placeholder_excel_template)
        sheet = wb['Summary']
        
        # Calculate aggregates
        total_spend = sample_campaign_df['Spend_USD'].sum()
        total_revenue = sample_campaign_df['Revenue_USD'].sum()
        total_clicks = sample_campaign_df['Clicks'].sum()
        
        # Replace placeholders
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if '{{Total_Spend}}' in cell.value:
                        cell.value = round(total_spend, 2)
                    elif '{{Total_Revenue}}' in cell.value:
                        cell.value = round(total_revenue, 2)
                    elif '{{Total_Clicks}}' in cell.value:
                        cell.value = total_clicks
        
        output_path = Path(temp_dir) / "placeholder_output.xlsx"
        wb.save(output_path)
        
        assert output_path.exists()
        
        # Verify placeholders were replaced
        wb_check = load_workbook(output_path)
        sheet_check = wb_check['Summary']
        assert sheet_check['B3'].value == round(total_spend, 2)
    
    def test_formula_generation(self, temp_dir):
        """Test generation of Excel formulas."""
        from openpyxl import Workbook, load_workbook
        
        wb = Workbook()
        ws = wb.active
        
        # Add data
        ws['A1'] = 'Impressions'
        ws['B1'] = 'Clicks'
        ws['C1'] = 'CTR'
        ws['A2'] = 1000
        ws['B2'] = 50
        ws['C2'] = '=IF(A2=0,0,B2/A2)'
        
        output_path = Path(temp_dir) / "formula_test.xlsx"
        wb.save(output_path)
        
        # Reload and check formula
        wb_check = load_workbook(output_path)
        ws_check = wb_check.active
        assert ws_check['C2'].value == '=IF(A2=0,0,B2/A2)'
    
    def test_preserve_formatting(self, sample_excel_template, temp_dir):
        """Test that formatting is preserved during report generation."""
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = load_workbook(sample_excel_template)
        sheet = wb['Daily Pacing']
        
        # Add formatting to header
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        output_path = Path(temp_dir) / "formatted_output.xlsx"
        wb.save(output_path)
        
        # Verify formatting preserved
        wb_check = load_workbook(output_path)
        sheet_check = wb_check['Daily Pacing']
        assert sheet_check['A1'].font.bold == True


# =============================================================================
# CSV REPORT GENERATION TESTS
# =============================================================================

class TestCSVReportGeneration:
    """Tests for CSV report generation."""
    
    def test_generate_csv_report(self, sample_campaign_df, temp_dir):
        """Test basic CSV report generation."""
        output_path = Path(temp_dir) / "output_report.csv"
        sample_campaign_df.to_csv(output_path, index=False)
        
        assert output_path.exists()
        
        # Verify data
        df_check = pd.read_csv(output_path)
        assert len(df_check) == len(sample_campaign_df)
        assert list(df_check.columns) == list(sample_campaign_df.columns)
    
    def test_csv_with_aggregation(self, sample_campaign_df, temp_dir):
        """Test CSV generation with aggregated data."""
        agg_df = sample_campaign_df.groupby('Platform').agg({
            'Impressions': 'sum',
            'Clicks': 'sum',
            'Spend_USD': 'sum',
            'Revenue_USD': 'sum'
        }).reset_index()
        
        output_path = Path(temp_dir) / "aggregated_report.csv"
        agg_df.to_csv(output_path, index=False)
        
        assert output_path.exists()
        
        df_check = pd.read_csv(output_path)
        assert len(df_check) == sample_campaign_df['Platform'].nunique()


# =============================================================================
# CACHE FUNCTIONALITY TESTS
# =============================================================================

class TestCacheFunctionality:
    """Tests for cache save/load functionality."""
    
    def test_save_to_cache(self, temp_dir, sample_campaign_df):
        """Test saving data to cache."""
        import pickle
        
        cache_dir = Path(temp_dir) / ".cache"
        cache_dir.mkdir(exist_ok=True)
        
        cache_data = {
            'data': sample_campaign_df,
            'filename': 'test_data.csv',
            'timestamp': datetime.now().isoformat()
        }
        
        cache_path = cache_dir / "data_df.pkl"
        with open(cache_path, 'wb') as f:
            pickle.dump(cache_data, f)
        
        assert cache_path.exists()
    
    def test_load_from_cache(self, temp_dir, sample_campaign_df):
        """Test loading data from cache."""
        import pickle
        
        cache_dir = Path(temp_dir) / ".cache"
        cache_dir.mkdir(exist_ok=True)
        
        # Save first
        cache_data = {
            'data': sample_campaign_df,
            'filename': 'test_data.csv',
            'timestamp': datetime.now().isoformat()
        }
        cache_path = cache_dir / "data_df.pkl"
        with open(cache_path, 'wb') as f:
            pickle.dump(cache_data, f)
        
        # Load
        with open(cache_path, 'rb') as f:
            loaded = pickle.load(f)
        
        assert loaded['filename'] == 'test_data.csv'
        assert len(loaded['data']) == len(sample_campaign_df)
    
    def test_cache_template_bytes(self, temp_dir, sample_excel_template):
        """Test caching template file as bytes."""
        import pickle
        
        cache_dir = Path(temp_dir) / ".cache"
        cache_dir.mkdir(exist_ok=True)
        
        # Read template as bytes
        with open(sample_excel_template, 'rb') as f:
            template_bytes = f.read()
        
        cache_data = {
            'data': template_bytes,
            'filename': 'template.xlsx',
            'timestamp': datetime.now().isoformat()
        }
        
        cache_path = cache_dir / "template.pkl"
        with open(cache_path, 'wb') as f:
            pickle.dump(cache_data, f)
        
        # Load and verify
        with open(cache_path, 'rb') as f:
            loaded = pickle.load(f)
        
        assert loaded['data'] == template_bytes
    
    def test_clear_cache(self, temp_dir):
        """Test clearing cache files."""
        import pickle
        
        cache_dir = Path(temp_dir) / ".cache"
        cache_dir.mkdir(exist_ok=True)
        
        # Create some cache files
        for name in ['file1.pkl', 'file2.pkl', 'file3.pkl']:
            with open(cache_dir / name, 'wb') as f:
                pickle.dump({'test': 'data'}, f)
        
        assert len(list(cache_dir.glob("*.pkl"))) == 3
        
        # Clear cache
        for cache_file in cache_dir.glob("*.pkl"):
            cache_file.unlink()
        
        assert len(list(cache_dir.glob("*.pkl"))) == 0


# =============================================================================
# UPDATE PACING TRACKER TESTS
# =============================================================================

class TestUpdatePacingTracker:
    """Tests for update_pacing_tracker.py functionality."""
    
    def test_update_pacing_tracker_import(self):
        """Test that update_pacing_tracker module can be imported."""
        try:
            from update_pacing_tracker import update_pacing_tracker
            assert callable(update_pacing_tracker)
        except ImportError as e:
            pytest.skip(f"update_pacing_tracker not available: {e}")
    
    def test_update_pacing_tracker_missing_csv(self, sample_excel_template):
        """Test error handling for missing CSV file."""
        try:
            from update_pacing_tracker import update_pacing_tracker
        except ImportError:
            pytest.skip("update_pacing_tracker not available")
        
        with pytest.raises(FileNotFoundError):
            update_pacing_tracker("/nonexistent/path.csv", sample_excel_template)
    
    def test_update_pacing_tracker_missing_template(self, sample_csv_file):
        """Test error handling for missing template file."""
        try:
            from update_pacing_tracker import update_pacing_tracker
        except ImportError:
            pytest.skip("update_pacing_tracker not available")
        
        with pytest.raises(FileNotFoundError):
            update_pacing_tracker(sample_csv_file, "/nonexistent/template.xlsx")
    
    def test_update_pacing_tracker_missing_columns(self, temp_dir, sample_excel_template):
        """Test error handling for missing required columns."""
        try:
            from update_pacing_tracker import update_pacing_tracker
        except ImportError:
            pytest.skip("update_pacing_tracker not available")
        
        # Create CSV with missing columns
        df = pd.DataFrame({'Date': ['2024-01-01'], 'Platform': ['Google']})
        csv_path = Path(temp_dir) / "incomplete.csv"
        df.to_csv(csv_path, index=False)
        
        with pytest.raises(KeyError):
            update_pacing_tracker(csv_path, sample_excel_template)
    
    def test_update_pacing_tracker_success(self, sample_csv_file, sample_excel_template, temp_dir):
        """Test successful pacing tracker update."""
        try:
            from update_pacing_tracker import update_pacing_tracker
        except ImportError:
            pytest.skip("update_pacing_tracker not available")
        
        output_path = Path(temp_dir) / "updated_tracker.xlsx"
        
        result = update_pacing_tracker(
            sample_csv_file,
            sample_excel_template,
            output_path
        )
        
        assert output_path.exists()
        assert 'records_processed' in result
        assert 'daily_records' in result
        assert 'weekly_records' in result
        assert result['records_processed'] == 20


# =============================================================================
# STREAMLIT REPORTING INTEGRATION TESTS
# =============================================================================

class TestStreamlitReportingIntegration:
    """Tests for streamlit_reporting.py integration."""
    
    def test_streamlit_reporting_exists(self):
        """Test that streamlit_reporting.py exists."""
        streamlit_path = PROJECT_ROOT / "streamlit_reporting.py"
        assert streamlit_path.exists(), "streamlit_reporting.py not found"
    
    def test_analyze_template_function_exists(self):
        """Test that analyze_template function is importable."""
        # We can't directly import streamlit modules without mocking st
        streamlit_path = PROJECT_ROOT / "streamlit_reporting.py"
        content = streamlit_path.read_text()
        
        assert 'def analyze_template(' in content
        assert 'def generate_excel_report(' in content
        assert 'def generate_csv_report(' in content
    
    def test_cache_functions_exist(self):
        """Test that cache functions are defined."""
        streamlit_path = PROJECT_ROOT / "streamlit_reporting.py"
        content = streamlit_path.read_text()
        
        assert 'def save_to_cache(' in content
        assert 'def load_from_cache(' in content
        assert 'def clear_cache(' in content
    
    def test_column_mode_support(self):
        """Test that column mode is supported."""
        streamlit_path = PROJECT_ROOT / "streamlit_reporting.py"
        content = streamlit_path.read_text()
        
        # Check for column mode indicators
        assert 'column' in content.lower()
        assert '_column_mappings' in content or 'template_columns' in content
    
    def test_placeholder_mode_support(self):
        """Test that placeholder mode is supported."""
        streamlit_path = PROJECT_ROOT / "streamlit_reporting.py"
        content = streamlit_path.read_text()
        
        assert "'mode': 'placeholder'" in content or '"mode": "placeholder"' in content
        assert 'placeholders' in content
    
    def test_template_structure_keys(self):
        """Test that template structure includes required keys."""
        streamlit_path = PROJECT_ROOT / "streamlit_reporting.py"
        content = streamlit_path.read_text()
        
        assert "'template_columns'" in content
        assert "'data_start_row'" in content
        assert "'sheet_count'" in content
        assert "'field_count'" in content


# =============================================================================
# COLUMN MATCHING TESTS
# =============================================================================

class TestColumnMatching:
    """Tests for column matching logic."""
    
    def test_exact_match(self):
        """Test exact column name matching."""
        def find_matching_column(placeholder, columns):
            clean = placeholder.strip('{}[]<>').lower().replace('_', ' ').replace('-', ' ')
            for col in columns:
                if col.lower() == clean:
                    return col
            for col in columns:
                if clean in col.lower() or col.lower() in clean:
                    return col
            return None
        
        columns = ['Date', 'Platform', 'Impressions', 'Clicks', 'Spend_USD']
        
        assert find_matching_column('Date', columns) == 'Date'
        assert find_matching_column('{{Date}}', columns) == 'Date'
        assert find_matching_column('[Platform]', columns) == 'Platform'
    
    def test_partial_match(self):
        """Test partial column name matching."""
        def find_matching_column(placeholder, columns):
            clean = placeholder.strip('{}[]<>').lower().replace('_', ' ').replace('-', ' ')
            for col in columns:
                if col.lower() == clean:
                    return col
            for col in columns:
                if clean in col.lower() or col.lower() in clean:
                    return col
            return None
        
        columns = ['Campaign_Name', 'AdGroup_Name', 'Total_Spend_USD']
        
        assert find_matching_column('Campaign', columns) == 'Campaign_Name'
        assert find_matching_column('Spend', columns) == 'Total_Spend_USD'
    
    def test_no_match(self):
        """Test handling of unmatched columns."""
        def find_matching_column(placeholder, columns):
            clean = placeholder.strip('{}[]<>').lower().replace('_', ' ').replace('-', ' ')
            for col in columns:
                if col.lower() == clean:
                    return col
            for col in columns:
                if clean in col.lower() or col.lower() in clean:
                    return col
            return None
        
        columns = ['Date', 'Platform', 'Impressions']
        
        assert find_matching_column('NonExistent', columns) is None


# =============================================================================
# DATA VALIDATION TESTS
# =============================================================================

class TestDataValidation:
    """Tests for data validation logic."""
    
    def test_validate_required_columns(self, sample_campaign_df):
        """Test validation of required columns."""
        required = {'Date', 'Platform', 'Impressions', 'Clicks'}
        missing = required - set(sample_campaign_df.columns)
        
        assert len(missing) == 0
    
    def test_validate_numeric_values(self, sample_campaign_df):
        """Test validation of numeric column values."""
        numeric_cols = ['Impressions', 'Clicks', 'Spend_USD', 'Conversions', 'Revenue_USD']
        
        for col in numeric_cols:
            assert (sample_campaign_df[col] >= 0).all(), f"{col} has negative values"
    
    def test_validate_date_range(self, sample_campaign_df):
        """Test validation of date range."""
        min_date = sample_campaign_df['Date'].min()
        max_date = sample_campaign_df['Date'].max()
        
        assert min_date <= max_date
        assert (max_date - min_date).days < 365  # Reasonable range
    
    def test_handle_null_values(self):
        """Test handling of null values."""
        df = pd.DataFrame({
            'Impressions': [1000, None, 500],
            'Clicks': [50, 25, None],
        })
        
        # Fill nulls with 0
        df_filled = df.fillna(0)
        assert df_filled['Impressions'].isna().sum() == 0
        assert df_filled['Clicks'].isna().sum() == 0


# =============================================================================
# EDGE CASES TESTS
# =============================================================================

class TestEdgeCases:
    """Tests for edge cases and error handling."""
    
    def test_empty_dataframe(self):
        """Test handling of empty DataFrame."""
        df = pd.DataFrame(columns=['Date', 'Platform', 'Impressions'])
        
        assert len(df) == 0
        assert df['Impressions'].sum() == 0
    
    def test_single_row_dataframe(self):
        """Test handling of single-row DataFrame."""
        df = pd.DataFrame({
            'Date': [date(2024, 1, 1)],
            'Platform': ['Google Ads'],
            'Impressions': [1000],
            'Clicks': [50],
        })
        
        assert len(df) == 1
        assert df['Impressions'].sum() == 1000
    
    def test_large_numbers(self):
        """Test handling of large numbers."""
        df = pd.DataFrame({
            'Impressions': [1_000_000_000],
            'Spend': [10_000_000.50],
        })
        
        assert df['Impressions'].iloc[0] == 1_000_000_000
        assert df['Spend'].iloc[0] == 10_000_000.50
    
    def test_special_characters_in_names(self):
        """Test handling of special characters in column/campaign names."""
        df = pd.DataFrame({
            'Campaign_Name': ['Campaign (Test)', 'Campaign & More', 'Campaign #1'],
            'Impressions': [100, 200, 300],
        })
        
        assert len(df) == 3
        assert df['Campaign_Name'].iloc[0] == 'Campaign (Test)'
    
    def test_unicode_characters(self):
        """Test handling of unicode characters."""
        df = pd.DataFrame({
            'Campaign_Name': ['キャンペーン', 'Campaña', 'Кампания'],
            'Impressions': [100, 200, 300],
        })
        
        assert len(df) == 3


# =============================================================================
# PERFORMANCE TESTS
# =============================================================================

class TestPerformance:
    """Tests for performance with larger datasets."""
    
    def test_large_dataset_aggregation(self):
        """Test aggregation performance with larger dataset."""
        # Create 10,000 row dataset
        df = pd.DataFrame({
            'Date': pd.date_range('2024-01-01', periods=10000, freq='H'),
            'Platform': np.random.choice(['Google', 'Meta', 'Snap'], 10000),
            'Impressions': np.random.randint(1000, 100000, 10000),
            'Clicks': np.random.randint(10, 1000, 10000),
            'Spend': np.random.uniform(10, 1000, 10000),
        })
        
        import time
        start = time.time()
        
        agg = df.groupby(['Platform']).agg({
            'Impressions': 'sum',
            'Clicks': 'sum',
            'Spend': 'sum',
        }).reset_index()
        
        elapsed = time.time() - start
        
        assert elapsed < 1.0  # Should complete in under 1 second
        assert len(agg) == 3


# =============================================================================
# RUN TESTS AND GENERATE COVERAGE REPORT
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
