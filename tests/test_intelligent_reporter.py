"""
Tests for Intelligent Reporting System

Comprehensive tests covering:
1. Multi-source data reading
2. Hybrid field mapping
3. Data transformation
4. Template analysis
5. Report generation
"""

import pytest
import pandas as pd
import numpy as np
from datetime import datetime, date
import tempfile
import shutil
from pathlib import Path
import sys

# Add project root
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Set dummy env var to avoid settings validation error
import os
os.environ.setdefault('OPENAI_API_KEY', 'dummy-key-for-testing')

# Now import
from src.reporting.intelligent_reporter import (
    IntelligentReportSystem,
    IntelligentDataReader,
    FieldMappingEngine,
    DataTransformationEngine,
    TemplateAnalyzer,
    IntelligentTemplateUpdater,
    FieldMapping,
    TemplateStructure,
    DataSourceConfig,
    ReportResult,
    FileAdapter,
    generate_report,
)


# =============================================================================
# FIXTURES
# =============================================================================

@pytest.fixture
def temp_dir():
    """Create temporary directory."""
    tmpdir = tempfile.mkdtemp()
    yield tmpdir
    shutil.rmtree(tmpdir, ignore_errors=True)


@pytest.fixture
def sample_campaign_df():
    """Sample campaign data."""
    return pd.DataFrame({
        'Date': pd.date_range('2024-07-01', periods=20, freq='D'),
        'Platform': ['Google Ads', 'Meta Ads', 'Snapchat', 'YouTube', 'Bing'] * 4,
        'Campaign_Name': [f'Campaign_{i}' for i in range(20)],
        'Impressions': np.random.randint(10000, 500000, 20),
        'Clicks': np.random.randint(100, 10000, 20),
        'Spend_USD': np.random.uniform(100, 5000, 20).round(2),
        'Conversions': np.random.randint(10, 500, 20),
        'Revenue_USD': np.random.uniform(500, 25000, 20).round(2),
    })


@pytest.fixture
def sample_csv_file(temp_dir, sample_campaign_df):
    """Create sample CSV file."""
    path = Path(temp_dir) / "campaign_data.csv"
    sample_campaign_df.to_csv(path, index=False)
    return path


@pytest.fixture
def sample_excel_file(temp_dir, sample_campaign_df):
    """Create sample Excel file."""
    path = Path(temp_dir) / "campaign_data.xlsx"
    sample_campaign_df.to_excel(path, index=False)
    return path


@pytest.fixture
def sample_template(temp_dir):
    """Create sample Excel template."""
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Daily Pacing sheet
    ws = wb.active
    ws.title = "Daily Pacing"
    headers = ['Date', 'Platform', 'Impressions', 'Clicks', 'CTR', 
               'Conversions', 'Spend', 'CPC', 'CPA', 'Revenue', 'ROAS']
    for col, header in enumerate(headers, 1):
        ws.cell(row=7, column=col, value=header)
    
    # Weekly Pacing sheet
    ws2 = wb.create_sheet("Weekly Pacing")
    weekly_headers = ['Week', 'Start Date', 'End Date', 'Impressions', 
                      'Clicks', 'Spend', 'Conversions', 'Revenue']
    for col, header in enumerate(weekly_headers, 1):
        ws2.cell(row=5, column=col, value=header)
    
    # Summary sheet
    ws3 = wb.create_sheet("Summary")
    ws3['A1'] = "Campaign Summary"
    ws3['A3'] = "Start Date:"
    ws3['B3'] = ""
    ws3['A4'] = "End Date:"
    ws3['B4'] = ""
    ws3['A5'] = "Total Spend:"
    ws3['B5'] = ""
    
    path = Path(temp_dir) / "template.xlsx"
    wb.save(path)
    return path


# =============================================================================
# DATA READER TESTS
# =============================================================================

class TestIntelligentDataReader:
    """Tests for IntelligentDataReader."""
    
    def test_read_csv(self, sample_csv_file):
        """Test reading CSV file."""
        reader = IntelligentDataReader()
        df = reader.read(sample_csv_file)
        
        assert len(df) == 20
        assert 'Date' in df.columns
        assert reader.metadata['row_count'] == 20
    
    def test_read_excel(self, sample_excel_file):
        """Test reading Excel file."""
        reader = IntelligentDataReader()
        df = reader.read(sample_excel_file)
        
        assert len(df) == 20
        assert 'Platform' in df.columns
    
    def test_read_dataframe_passthrough(self, sample_campaign_df):
        """Test DataFrame passthrough."""
        reader = IntelligentDataReader()
        df = reader.read(sample_campaign_df)
        
        assert len(df) == len(sample_campaign_df)
        assert list(df.columns) == list(sample_campaign_df.columns)
    
    def test_quality_report(self, sample_csv_file):
        """Test quality report generation."""
        reader = IntelligentDataReader()
        reader.read(sample_csv_file)
        
        assert reader.quality_report is not None
        assert 'total_rows' in reader.quality_report
        assert 'column_stats' in reader.quality_report


# =============================================================================
# FIELD MAPPING ENGINE TESTS
# =============================================================================

class TestFieldMappingEngine:
    """Tests for FieldMappingEngine."""
    
    def test_exact_mapping(self):
        """Test exact field matching."""
        mapper = FieldMappingEngine()
        mappings = mapper.map_fields(
            source_columns=['date', 'platform', 'impressions', 'clicks'],
            target_columns=['date', 'platform', 'impressions', 'clicks']
        )
        
        assert mappings['date'].target_field == 'date'
        assert mappings['date'].confidence == 1.0
        assert mappings['date'].method == 'exact'
    
    def test_alias_mapping(self):
        """Test alias-based mapping."""
        mapper = FieldMappingEngine()
        mappings = mapper.map_fields(
            source_columns=['cost', 'impr', 'conv'],
            target_columns=['spend', 'impressions', 'conversions']
        )
        
        assert mappings['cost'].target_field == 'spend'
        assert mappings['impr'].target_field == 'impressions'
        assert mappings['conv'].target_field == 'conversions'
    
    def test_fuzzy_mapping(self):
        """Test fuzzy matching."""
        mapper = FieldMappingEngine(confidence_threshold=0.5)
        mappings = mapper.map_fields(
            source_columns=['total_impressions', 'click_count'],
            target_columns=['impressions', 'clicks']
        )
        
        assert mappings['total_impressions'].target_field == 'impressions'
        assert mappings['click_count'].target_field == 'clicks'
    
    def test_manual_overrides(self):
        """Test manual mapping overrides."""
        mapper = FieldMappingEngine()
        mappings = mapper.map_fields(
            source_columns=['weird_column', 'another_weird'],
            target_columns=['spend', 'revenue'],
            manual_overrides={'weird_column': 'spend'}
        )
        
        assert mappings['weird_column'].target_field == 'spend'
        assert mappings['weird_column'].method == 'manual'
    
    def test_unmatched_columns(self):
        """Test handling of unmatched columns."""
        mapper = FieldMappingEngine(confidence_threshold=0.9)
        mappings = mapper.map_fields(
            source_columns=['xyz_unknown', 'abc_random'],
            target_columns=['spend', 'revenue']
        )
        
        assert mappings['xyz_unknown'].target_field is None
        assert mappings['xyz_unknown'].method == 'unmatched'


# =============================================================================
# DATA TRANSFORMATION TESTS
# =============================================================================

class TestDataTransformationEngine:
    """Tests for DataTransformationEngine."""
    
    def test_transform_basic(self, sample_campaign_df):
        """Test basic transformation."""
        transformer = DataTransformationEngine()
        
        # Create simple mappings
        mappings = {
            'Date': FieldMapping('Date', 'date', 1.0, 'exact'),
            'Platform': FieldMapping('Platform', 'platform', 1.0, 'exact'),
            'Impressions': FieldMapping('Impressions', 'impressions', 1.0, 'exact'),
            'Clicks': FieldMapping('Clicks', 'clicks', 1.0, 'exact'),
            'Spend_USD': FieldMapping('Spend_USD', 'spend', 1.0, 'exact'),
            'Conversions': FieldMapping('Conversions', 'conversions', 1.0, 'exact'),
            'Revenue_USD': FieldMapping('Revenue_USD', 'revenue', 1.0, 'exact'),
        }
        
        result = transformer.transform(sample_campaign_df, mappings)
        
        assert 'raw' in result
        assert 'daily' in result
        assert 'weekly' in result
        assert 'summary' in result
    
    def test_calculated_metrics(self, sample_campaign_df):
        """Test calculated metrics."""
        transformer = DataTransformationEngine()
        
        mappings = {
            'Date': FieldMapping('Date', 'date', 1.0, 'exact'),
            'Impressions': FieldMapping('Impressions', 'impressions', 1.0, 'exact'),
            'Clicks': FieldMapping('Clicks', 'clicks', 1.0, 'exact'),
            'Spend_USD': FieldMapping('Spend_USD', 'spend', 1.0, 'exact'),
            'Conversions': FieldMapping('Conversions', 'conversions', 1.0, 'exact'),
            'Revenue_USD': FieldMapping('Revenue_USD', 'revenue', 1.0, 'exact'),
        }
        
        result = transformer.transform(sample_campaign_df, mappings)
        daily = result['daily']  # Calculated metrics are in aggregated data
        
        assert 'ctr' in daily.columns
        assert 'cpc' in daily.columns
        assert 'cpa' in daily.columns
        assert 'roas' in daily.columns
    
    def test_summary_generation(self, sample_campaign_df):
        """Test summary statistics."""
        transformer = DataTransformationEngine()
        
        mappings = {
            'Date': FieldMapping('Date', 'date', 1.0, 'exact'),
            'Impressions': FieldMapping('Impressions', 'impressions', 1.0, 'exact'),
            'Spend_USD': FieldMapping('Spend_USD', 'spend', 1.0, 'exact'),
        }
        
        result = transformer.transform(sample_campaign_df, mappings)
        summary = result['summary']
        
        assert 'total_rows' in summary
        assert 'date_range' in summary
        assert 'totals' in summary
        assert summary['totals']['impressions'] == sample_campaign_df['Impressions'].sum()


# =============================================================================
# TEMPLATE ANALYZER TESTS
# =============================================================================

class TestTemplateAnalyzer:
    """Tests for TemplateAnalyzer."""
    
    def test_analyze_template(self, sample_template):
        """Test template analysis."""
        analyzer = TemplateAnalyzer(sample_template)
        structures = analyzer.analyze()
        
        assert 'Daily Pacing' in structures
        assert 'Weekly Pacing' in structures
        assert 'Summary' in structures
    
    def test_detect_header_row(self, sample_template):
        """Test header row detection."""
        analyzer = TemplateAnalyzer(sample_template)
        structures = analyzer.analyze()
        
        daily = structures['Daily Pacing']
        assert daily.header_row == 7
        assert daily.data_start_row == 8
    
    def test_extract_columns(self, sample_template):
        """Test column extraction."""
        analyzer = TemplateAnalyzer(sample_template)
        structures = analyzer.analyze()
        
        daily = structures['Daily Pacing']
        assert 'Date' in daily.columns
        assert 'Platform' in daily.columns
        assert 'Impressions' in daily.columns
    
    def test_detect_structure_type(self, sample_template):
        """Test structure type detection."""
        analyzer = TemplateAnalyzer(sample_template)
        structures = analyzer.analyze()
        
        assert structures['Daily Pacing'].structure_type == 'data_table'
        assert structures['Summary'].structure_type == 'summary'
    
    def test_detect_aggregation_type(self, sample_template):
        """Test aggregation type detection."""
        analyzer = TemplateAnalyzer(sample_template)
        structures = analyzer.analyze()
        
        assert structures['Daily Pacing'].aggregation_type == 'daily'
        assert structures['Weekly Pacing'].aggregation_type == 'weekly'


# =============================================================================
# FULL SYSTEM TESTS
# =============================================================================

class TestIntelligentReportSystem:
    """Tests for the full IntelligentReportSystem."""
    
    def test_generate_report_csv(self, sample_csv_file, sample_template, temp_dir):
        """Test report generation from CSV."""
        output_path = Path(temp_dir) / "output.xlsx"
        
        system = IntelligentReportSystem()
        result = system.generate_report(
            data_source=sample_csv_file,
            template_path=sample_template,
            output_path=output_path
        )
        
        assert result.success
        assert output_path.exists()
        assert len(result.sheets_updated) > 0
        assert result.rows_written > 0
    
    def test_generate_report_dataframe(self, sample_campaign_df, sample_template, temp_dir):
        """Test report generation from DataFrame."""
        output_path = Path(temp_dir) / "output.xlsx"
        
        system = IntelligentReportSystem()
        result = system.generate_report(
            data_source=sample_campaign_df,
            template_path=sample_template,
            output_path=output_path
        )
        
        assert result.success
        assert output_path.exists()
    
    def test_generate_report_with_hints(self, sample_csv_file, sample_template, temp_dir):
        """Test report generation with mapping hints."""
        output_path = Path(temp_dir) / "output.xlsx"
        
        system = IntelligentReportSystem()
        result = system.generate_report(
            data_source=sample_csv_file,
            template_path=sample_template,
            output_path=output_path,
            mapping_hints={'Spend_USD': 'spend'}
        )
        
        assert result.success
        assert 'Spend_USD' in result.mappings_used
        assert result.mappings_used['Spend_USD'].method == 'manual'
    
    def test_convenience_function(self, sample_csv_file, sample_template, temp_dir):
        """Test convenience function."""
        output_path = Path(temp_dir) / "output.xlsx"
        
        result = generate_report(
            sample_csv_file,
            sample_template,
            output_path
        )
        
        assert result.success


# =============================================================================
# EDGE CASES
# =============================================================================

class TestEdgeCases:
    """Tests for edge cases."""
    
    def test_empty_dataframe(self, sample_template, temp_dir):
        """Test handling of empty DataFrame."""
        output_path = Path(temp_dir) / "output.xlsx"
        empty_df = pd.DataFrame()
        
        system = IntelligentReportSystem()
        result = system.generate_report(
            data_source=empty_df,
            template_path=sample_template,
            output_path=output_path
        )
        
        # Should handle gracefully
        assert output_path.exists() or len(result.warnings) > 0
    
    def test_missing_columns(self, sample_template, temp_dir):
        """Test handling of data with missing expected columns."""
        output_path = Path(temp_dir) / "output.xlsx"
        minimal_df = pd.DataFrame({
            'Date': pd.date_range('2024-01-01', periods=5),
            'Value': [1, 2, 3, 4, 5]
        })
        
        system = IntelligentReportSystem()
        result = system.generate_report(
            data_source=minimal_df,
            template_path=sample_template,
            output_path=output_path
        )
        
        # Should handle gracefully
        assert output_path.exists()
    
    def test_weird_column_names(self, sample_template, temp_dir):
        """Test handling of unusual column names."""
        output_path = Path(temp_dir) / "output.xlsx"
        weird_df = pd.DataFrame({
            'report_date': pd.date_range('2024-01-01', periods=5),
            'ad_network': ['Google'] * 5,
            'total_impr': [1000] * 5,
            'user_clicks': [50] * 5,
            'campaign_cost': [100.0] * 5,
        })
        
        system = IntelligentReportSystem()
        result = system.generate_report(
            data_source=weird_df,
            template_path=sample_template,
            output_path=output_path
        )
        
        assert result.success
        # Should map weird names to standard ones
        assert any(m.target_field == 'date' for m in result.mappings_used.values())
        assert any(m.target_field == 'impressions' for m in result.mappings_used.values())


# =============================================================================
# FILE ADAPTER TESTS
# =============================================================================

class TestFileAdapter:
    """Tests for FileAdapter."""
    
    def test_validate_csv(self, sample_csv_file):
        """Test CSV validation."""
        adapter = FileAdapter()
        config = DataSourceConfig(source_type='csv', path_or_connection=str(sample_csv_file))
        
        assert adapter.validate(config)
    
    def test_validate_nonexistent(self, temp_dir):
        """Test validation of non-existent file."""
        adapter = FileAdapter()
        config = DataSourceConfig(source_type='csv', path_or_connection=str(Path(temp_dir) / "nonexistent.csv"))
        
        assert not adapter.validate(config)
    
    def test_read_with_encoding(self, temp_dir):
        """Test reading with different encodings."""
        # Create file with latin-1 encoding
        path = Path(temp_dir) / "latin1.csv"
        df = pd.DataFrame({'col': ['café', 'naïve']})
        df.to_csv(path, index=False, encoding='latin-1')
        
        adapter = FileAdapter()
        config = DataSourceConfig(source_type='csv', path_or_connection=str(path))
        result = adapter.read(config)
        
        assert len(result) == 2


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
