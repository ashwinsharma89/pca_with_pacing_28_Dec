
import sys
from pathlib import Path
import pandas as pd
from datetime import datetime
from loguru import logger

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.agents.pacing_report_agent import PacingReportAgent, AggregationLevel

def create_sample_data():
    """Create sample campaign data with multiple platforms and channels."""
    data = {
        'Date': pd.date_range(start='2024-12-01', periods=30).tolist() * 3,
        'Platform': (['Google Ads'] * 30) + (['Meta Ads'] * 30) + (['LinkedIn Ads'] * 30),
        'Campaign': (['Brand search'] * 30) + (['Retargeting'] * 30) + (['ABM'] * 30),
        'Impressions': [1000] * 90,
        'Clicks': [100] * 90,
        'Spend_USD': [500.0] * 90,
        'Conversions': [5] * 90,
        'Revenue_USD': [1500.0] * 90
    }
    return pd.DataFrame(data)

def test_pacing_report_agent():
    # 1. Setup Agent
    agent = PacingReportAgent()
    
    # 2. Prepare Template
    # We'll use the production budget pacing template which has the 'Pivot Analysis' sheet
    templates_dir = Path("reports/pacing/templates")
    template_path = templates_dir / "template_20251226_211223_pca_budget_pacing_template.xlsx"
    
    if not template_path.exists():
        logger.error(f"Template not found: {template_path}")
        return

    # 3. Mock data fetching
    df = create_sample_data()
    # Mock the fetch_campaign_data method to return our local dataframe
    agent.fetch_campaign_data = lambda start, end, filters: df
    
    # 4. Generate Report
    logger.info("Generating test report...")
    output_filename = "test_pacing_report_persistence.xlsx"
    result = agent.generate_report(
        template_path=template_path,
        start_date="2024-12-01",
        end_date="2024-12-30",
        aggregation=AggregationLevel.DAILY,
        output_filename=output_filename
    )
    
    if result.get("success"):
        logger.success(f"Report generated successfully: {result['output_file']}")
        
        # 5. Verify Formula Persistence (Manual inspection or code-based check)
        import openpyxl
        wb = openpyxl.load_workbook(result['output_file'])
        if 'Pivot Analysis' in wb.sheetnames:
            sheet = wb['Pivot Analysis']
            logger.info("Inspecting 'Pivot Analysis' sheet for formulas...")
            # Check row 14 or others for formulas
            for row_idx in range(1, 40):
                cell = sheet.cell(row_idx, 2) # Column B
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    logger.info(f"Found formula at B{row_idx}: {cell.value}")
                elif cell.value:
                    logger.info(f"Value at B{row_idx}: {cell.value} (Type: {type(cell.value)})")
        else:
            logger.warning("'Pivot Analysis' sheet not found in the output.")
    else:
        logger.error(f"Report generation failed: {result.get('error')}")

if __name__ == "__main__":
    test_pacing_report_agent()
