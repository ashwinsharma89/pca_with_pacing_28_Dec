import reflex as rx
from typing import Dict, List, Any, Optional
from .predictive import PredictiveState
import pandas as pd
import io

# Optional connectors – lazy import to avoid hard dependency
try:
    import snowflake.connector as snowflake
except ImportError:
    snowflake = None
try:
    import databricks.sql as databricks_sql
except ImportError:
    databricks_sql = None
try:
    from google.cloud import bigquery
except ImportError:
    bigquery = None
try:
    import boto3
except ImportError:
    boto3 = None

try:
    from src.reporting import IntelligentReportSystem as ReportGenerator
    BACKEND_AVAILABLE = True
except ImportError:
    BACKEND_AVAILABLE = False
    print("Warning: Reporting backend modules not found.")

class ReportingState(PredictiveState):
    """State for Automated Reporting, now supporting file uploads and external data sources."""

    # --- Template handling (unchanged) ---
    template_filename: str = ""
    template_file_content: Optional[bytes] = None
    template_structure: Dict[str, Any] = {}

    # --- Data handling ---
    data_filename: str = ""
    data_file_content: Optional[bytes] = None
    data_structure: Dict[str, Any] = {}
    data_source_type: str = ""  # "upload", "snowflake", "databricks", "bigquery", "s3"
    data_source_config: Dict[str, Any] = {}

    # Mapping & generation (unchanged)
    mapping_config: Dict[str, Any] = {}
    generated_report_path: str = ""
    is_generating: bool = False
    active_step: int = 0  # 0: Upload, 1: Mapping, 2: Download

    # ---------------------------------------------------------------------
    # Template upload (existing logic retained)
    async def handle_template_upload(self, files: List[rx.UploadFile]):
        if not files:
            return
        file = files[0]
        self.template_filename = file.filename
        content = await file.read()
        self.template_file_content = content
        self.analyze_template_structure(file.filename, content)
        if self.template_structure:
            self.active_step = 1

    # ---------------------------------------------------------------------
    # Data upload handling (CSV/XLSX)
    async def handle_data_upload(self, files: List[rx.UploadFile]):
        """Handle upload of a data file (CSV or XLSX)."""
        if not files:
            return
        file = files[0]
        self.data_filename = file.filename
        content = await file.read()
        self.data_file_content = content
        self.data_source_type = "upload"
        # Simple analysis using pandas
        try:
            if file.filename.lower().endswith('.csv'):
                df = pd.read_csv(io.BytesIO(content))
            elif file.filename.lower().endswith('.xlsx'):
                df = pd.read_excel(io.BytesIO(content))
            else:
                self.data_structure = {"error": "Unsupported file type"}
                return
            self.data_structure = {
                "rows": len(df),
                "columns": list(df.columns),
                "preview": df.head(5).to_dict(orient='records')
            }
        except Exception as e:
            self.data_structure = {"error": str(e)}
        # Move to next step if template already uploaded
        if self.template_structure:
            self.active_step = 1

    def set_data_source_type(self, source_type: str):
        """Select external data source type and reset related fields."""
        self.data_source_type = source_type
        self.data_source_config = {}
        self.data_structure = {}
        self.data_file_content = None
        self.data_filename = ""
        # Reset step to allow configuration
        self.active_step = 0

    def set_data_source(self, source_type: str, config: Dict[str, Any]):
        """Store external source configuration and attempt a preview fetch.
        Minimal implementation; actual connectors are lazily imported.
        """
        self.data_source_type = source_type
        self.data_source_config = config
        try:
            if source_type == "snowflake" and snowflake:
                conn = snowflake.connect(
                    user=config.get('user'),
                    password=config.get('password'),
                    account=config.get('account'),
                    warehouse=config.get('warehouse'),
                    database=config.get('database'),
                    schema=config.get('schema')
                )
                cur = conn.cursor()
                cur.execute(config.get('query'))
                rows = cur.fetchmany(5)
                columns = [desc[0] for desc in cur.description]
                self.data_structure = {"preview": [dict(zip(columns, r)) for r in rows]}
                conn.close()
            elif source_type == "databricks" and databricks_sql:
                conn = databricks_sql.connect(**config)
                cur = conn.cursor()
                cur.execute(config.get('query'))
                rows = cur.fetchmany(5)
                columns = [desc[0] for desc in cur.description]
                self.data_structure = {"preview": [dict(zip(columns, r)) for r in rows]}
                conn.close()
            elif source_type == "bigquery" and bigquery:
                client = bigquery.Client()
                query_job = client.query(config.get('query'))
                df = query_job.to_dataframe().head()
                self.data_structure = {
                    "rows": len(df),
                    "columns": list(df.columns),
                    "preview": df.to_dict(orient='records')
                }
            elif source_type == "s3" and boto3:
                s3 = boto3.client('s3')
                obj = s3.get_object(Bucket=config.get('bucket'), Key=config.get('key'))
                body = obj['Body'].read()
                df = pd.read_csv(io.BytesIO(body))
                self.data_structure = {
                    "rows": len(df),
                    "columns": list(df.columns),
                    "preview": df.head(5).to_dict(orient='records')
                }
            else:
                self.data_structure = {"error": "Unsupported source type or missing dependency"}
        except Exception as e:
            self.data_structure = {"error": str(e)}
        self.active_step = 1

    def analyze_template_structure(self, filename: str, content: bytes):
        """Analyze template to find placeholders."""
        # Simplified logic for MVP - assumes XLSX with {{placeholders}}
        if filename.endswith(".xlsx"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                placeholders = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if isinstance(cell, str) and "{{" in cell and "}}" in cell:
                                # Extract content between braces
                                import re
                                matches = re.findall(r"\{\{(.*?)\}\}", cell)
                                for m in matches:
                                    placeholders.append(m.strip())
                
                # Remove duplicates
                placeholders = list(set(placeholders))
                
                self.template_structure = {
                    "type": "xlsx",
                    "placeholders": placeholders,
                    "sheet_count": len(wb.sheetnames)
                }
            except Exception as e:
                print(f"Error analyzing template: {e}")
                self.template_structure = {"error": str(e)}

    def set_mapping(self, placeholder: str, column: str):
        """Set mapping for a placeholder."""
        self.mapping_config[placeholder] = column

    def generate_report(self):
        """Generate the final report."""
        if not self.template_file_content:
             yield rx.window_alert("No template uploaded.")
             return
             
        self.is_generating = True
        yield
        
        try:
            # MVP Generation Logic
            # In a real app we'd call src.reporting.ReportGenerator
            
            # Simulated delay
            import time
            time.sleep(1)
            
            self.active_step = 2
            self.generated_report_path = "report_generated.xlsx" # Mock
            
        except Exception as e:
            yield rx.window_alert(f"Generation Error: {e}")
        finally:
            self.is_generating = False
            
    def reset_reporting(self):
        # Reset all reporting‑related fields
        self.template_filename = ""
        self.template_file_content = None
        self.template_structure = {}
        self.data_filename = ""
        self.data_file_content = None
        self.data_structure = {}
        self.data_source_type = ""
        self.data_source_config = {}
        self.mapping_config = {}
        self.generated_report_path = ""
        self.active_step = 0
        self.is_generating = False
