"""Create a comprehensive test template with day-level tracking."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime, timedelta

# Create templates directory
Path('templates').mkdir(exist_ok=True)

# Create workbook
wb = openpyxl.Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ==================== SHEET 1: DAILY TRACKING ====================
ws_daily = wb.create_sheet('Daily Tracking')

# Title
ws_daily['A1'] = 'Daily Campaign Performance Tracker'
ws_daily['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_daily['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_daily.merge_cells('A1:J1')

# Summary Section
ws_daily['A3'] = 'SUMMARY METRICS'
ws_daily['A3'].font = Font(size=12, bold=True)
ws_daily['A3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws_daily.merge_cells('A3:D3')

ws_daily['A4'] = 'Total Spend:'
ws_daily['B4'] = '{{Total_Spend}}'
ws_daily['C4'] = 'Total Revenue:'
ws_daily['D4'] = '{{Total_Revenue}}'

ws_daily['A5'] = 'Overall ROAS:'
ws_daily['B5'] = '{{Overall_ROAS}}'
ws_daily['C5'] = 'Avg CTR:'
ws_daily['D5'] = '{{Avg_CTR}}%'

# Daily Data Table
ws_daily['A7'] = 'DAILY PERFORMANCE DATA'
ws_daily['A7'].font = Font(size=12, bold=True)
ws_daily['A7'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws_daily.merge_cells('A7:J7')

# Table Headers
headers = ['Date', 'Campaign', 'Spend', 'Impressions', 'Clicks', 'Conversions', 'Revenue', 'CTR', 'CPC', 'ROAS']
for col_idx, header in enumerate(headers, 1):
    cell = ws_daily.cell(row=8, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# Add sample date rows (30 days)
start_date = datetime.now() - timedelta(days=29)
for day in range(30):
    row = 9 + day
    current_date = start_date + timedelta(days=day)
    ws_daily.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))

# Set column widths
ws_daily.column_dimensions['A'].width = 12
ws_daily.column_dimensions['B'].width = 20
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws_daily.column_dimensions[col].width = 12

# ==================== SHEET 2: WEEKLY SUMMARY ====================
ws_weekly = wb.create_sheet('Weekly Summary')

# Title
ws_weekly['A1'] = 'Weekly Performance Summary'
ws_weekly['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_weekly['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws_weekly.merge_cells('A1:H1')

# Weekly Metrics
ws_weekly['A3'] = 'Week'
ws_weekly['B3'] = 'Total Spend'
ws_weekly['C3'] = 'Total Revenue'
ws_weekly['D3'] = 'Clicks'
ws_weekly['E3'] = 'Conversions'
ws_weekly['F3'] = 'ROAS'
ws_weekly['G3'] = 'CPC'
ws_weekly['H3'] = 'Status'

for col in range(1, 9):
    cell = ws_weekly.cell(row=3, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# Add 4 weeks
for week in range(1, 5):
    ws_weekly.cell(row=3 + week, column=1, value=f"Week {week}")

# Set column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_weekly.column_dimensions[col].width = 15

# ==================== SHEET 3: CAMPAIGN OVERVIEW ====================
ws_overview = wb.create_sheet('Campaign Overview')

# Title
ws_overview['A1'] = 'Campaign Overview Dashboard'
ws_overview['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_overview['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
ws_overview.merge_cells('A1:F1')

# Key Metrics
ws_overview['A3'] = 'KEY PERFORMANCE INDICATORS'
ws_overview['A3'].font = Font(size=12, bold=True)
ws_overview.merge_cells('A3:B3')

metrics = [
    ('Total Investment:', '{{Total_Spend}}'),
    ('Total Return:', '{{Total_Revenue}}'),
    ('Return on Ad Spend:', '{{Overall_ROAS}}x'),
    ('Average CPC:', '{{Avg_CPC}}'),
    ('Average CTR:', '{{Avg_CTR}}%'),
    ('Conversion Rate:', '{{Conversion_Rate}}%'),
]

for idx, (label, value) in enumerate(metrics, 4):
    ws_overview.cell(row=idx, column=1, value=label).font = Font(bold=True)
    ws_overview.cell(row=idx, column=2, value=value)

# Budget Tracking
ws_overview['A11'] = 'BUDGET TRACKING'
ws_overview['A11'].font = Font(size=12, bold=True)
ws_overview.merge_cells('A11:B11')

budget_metrics = [
    ('Total Budget:', '{{Budget_Total}}'),
    ('Budget Spent:', '{{Budget_Spent}}'),
    ('Budget Remaining:', '{{Budget_Remaining}}'),
    ('Utilization:', '{{Budget_Spent_Pct}}%'),
]

for idx, (label, value) in enumerate(budget_metrics, 12):
    ws_overview.cell(row=idx, column=1, value=label).font = Font(bold=True)
    ws_overview.cell(row=idx, column=2, value=value)

# Top Campaigns
ws_overview['D3'] = 'TOP PERFORMING CAMPAIGNS'
ws_overview['D3'].font = Font(size=12, bold=True)
ws_overview.merge_cells('D3:F3')

for i in range(1, 4):
    row = 3 + i
    ws_overview.cell(row=row, column=4, value=f"#{i}: {{{{Campaign{i}_Name}}}}")
    ws_overview.cell(row=row, column=5, value=f"{{{{Campaign{i}_Spend}}}}")
    ws_overview.cell(row=row, column=6, value=f"{{{{Campaign{i}_ROAS}}}}x")

# Set column widths
ws_overview.column_dimensions['A'].width = 20
ws_overview.column_dimensions['B'].width = 15
ws_overview.column_dimensions['D'].width = 25
ws_overview.column_dimensions['E'].width = 12
ws_overview.column_dimensions['F'].width = 10

# ==================== SHEET 4: PACING ANALYSIS ====================
ws_pacing = wb.create_sheet('Pacing Analysis')

# Title
ws_pacing['A1'] = 'Campaign Pacing Analysis'
ws_pacing['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_pacing['A1'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
ws_pacing.merge_cells('A1:G1')

# Pacing Table
headers_pacing = ['Campaign', 'Budget', 'Spent', 'Remaining', 'Days Left', 'Pacing %', 'Status']
for col_idx, header in enumerate(headers_pacing, 1):
    cell = ws_pacing.cell(row=3, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# Add 5 campaign rows
for row in range(4, 9):
    ws_pacing.cell(row=row, column=1, value=f"Campaign {row-3}")

# Set column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_pacing.column_dimensions[col].width = 15

# Save workbook
output_path = 'templates/comprehensive_test_template.xlsx'
wb.save(output_path)

print(f"[SUCCESS] Test template created: {output_path}")
print("\nTemplate Features:")
print("  - Daily Tracking sheet with 30-day table")
print("  - Weekly Summary sheet")
print("  - Campaign Overview dashboard")
print("  - Pacing Analysis sheet")
print("  - Multiple placeholder types")
print("  - Summary metrics")
print("  - Data tables with headers")
print("\nReady for testing!")
