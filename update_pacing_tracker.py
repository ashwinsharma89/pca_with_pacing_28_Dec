"""Campaign Pacing Tracker Auto-Update Script.

This utility ingests a normalized multi-platform reporting CSV and updates
an Excel pacing tracker template (daily + weekly tabs).
It mirrors the workflow described in the AI Agent guide so the PCA Agent
can automate reporting across CSV / DB inputs and generate XLSX / PPTX outputs.
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from openpyxl import load_workbook


def _safe_number(value):
    try:
        return float(value)
    except Exception:
        return 0.0


def update_pacing_tracker(
    csv_path: str | Path,
    excel_path: str | Path,
    output_path: Optional[str | Path] = None,
) -> Dict:
    """Update pacing tracker Excel file with data from CSV."""

    csv_path = Path(csv_path)
    excel_path = Path(excel_path)

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel template not found: {excel_path}")

    df = pd.read_csv(csv_path)
    required = {
        "Date",
        "Platform",
        "Impressions",
        "Clicks",
        "Spend_USD",
        "Conversions",
        "Revenue_USD",
    }
    missing = required - set(df.columns)
    if missing:
        raise KeyError(f"Missing columns: {sorted(missing)}")

    df["Date"] = pd.to_datetime(df["Date"])  # ensure datetime

    wb = load_workbook(excel_path)
    daily_sheet = wb["Daily Pacing"]
    weekly_sheet = wb["Weekly Pacing"]

    start_date = df["Date"].min()
    end_date = df["Date"].max()
    total_spend = _safe_number(df["Spend_USD"].sum())
    total_revenue = _safe_number(df["Revenue_USD"].sum())
    platforms = sorted(df["Platform"].astype(str).unique())
    platform_str = " / ".join(platforms)

    # campaign info
    daily_sheet["B3"] = "Multi-Platform Campaign Report"
    daily_sheet["F3"] = platform_str
    daily_sheet["J3"] = f"Multi_Campaign_{datetime.now():%Y%m%d_%H%M%S}"
    daily_sheet["B4"] = start_date
    daily_sheet["F4"] = end_date
    daily_sheet["J4"] = total_spend

    # daily aggregation
    daily_agg = (
        df.groupby(["Date", "Platform"], as_index=False)[
            ["Impressions", "Clicks", "Spend_USD", "Conversions", "Revenue_USD"]
        ]
        .sum()
        .sort_values("Date")
    )

    # clear rows 8-120
    for row in range(8, 121):
        for col in range(1, 15):
            daily_sheet.cell(row=row, column=col).value = None

    for idx, row_data in enumerate(daily_agg.itertuples(index=False), start=8):
        if idx > 200:
            break
        daily_sheet.cell(row=idx, column=1, value=row_data.Date)
        daily_sheet.cell(row=idx, column=2, value=row_data.Platform)
        daily_sheet.cell(row=idx, column=3, value=row_data.Impressions)
        daily_sheet.cell(row=idx, column=4, value=row_data.Clicks)
        daily_sheet.cell(row=idx, column=5, value=f"=IF(C{idx}=0,0,D{idx}/C{idx})")
        daily_sheet.cell(row=idx, column=6, value=row_data.Conversions)
        daily_sheet.cell(row=idx, column=7, value=f"=IF(D{idx}=0,0,F{idx}/D{idx})")
        daily_sheet.cell(row=idx, column=8, value=row_data.Spend_USD)
        daily_sheet.cell(row=idx, column=9, value=f"=IF(C{idx}=0,0,(H{idx}/C{idx})*1000)")
        daily_sheet.cell(row=idx, column=10, value=f"=IF(D{idx}=0,0,H{idx}/D{idx})")
        daily_sheet.cell(row=idx, column=11, value=f"=IF(F{idx}=0,0,H{idx}/F{idx})")
        daily_sheet.cell(row=idx, column=12, value=f"=IF(H{idx}=0,0,M{idx}/H{idx})")
        daily_sheet.cell(row=idx, column=13, value=row_data.Revenue_USD)

    last_daily = min(7 + len(daily_agg), 200)
    totals_row = last_daily + 2
    daily_sheet.cell(row=totals_row, column=1, value="TOTALS")
    for col, letter in enumerate("ABCDEFGHIJKLM", start=1):
        target_col = col
        cell = daily_sheet.cell(row=totals_row, column=target_col)
        if target_col in {3, 4, 6, 8, 13}:
            cell.value = f"=SUM({letter}8:{letter}{last_daily})"
        elif target_col == 5:
            cell.value = f"=IF(C{totals_row}=0,0,D{totals_row}/C{totals_row})"
        elif target_col == 7:
            cell.value = f"=IF(D{totals_row}=0,0,F{totals_row}/D{totals_row})"
        elif target_col == 9:
            cell.value = f"=IF(C{totals_row}=0,0,(H{totals_row}/C{totals_row})*1000)"
        elif target_col == 10:
            cell.value = f"=IF(D{totals_row}=0,0,H{totals_row}/D{totals_row})"
        elif target_col == 11:
            cell.value = f"=IF(F{totals_row}=0,0,H{totals_row}/F{totals_row})"
        elif target_col == 12:
            cell.value = f"=IF(H{totals_row}=0,0,M{totals_row}/H{totals_row})"

    # weekly aggregation
    df["Week"] = df["Date"].dt.isocalendar().week
    df["Year"] = df["Date"].dt.isocalendar().year
    weekly_agg = (
        df.groupby(["Year", "Week"], as_index=False)
        .agg(
            Week_Start=("Date", "min"),
            Week_End=("Date", "max"),
            Impressions=("Impressions", "sum"),
            Clicks=("Clicks", "sum"),
            Spend=("Spend_USD", "sum"),
            Conversions=("Conversions", "sum"),
            Revenue=("Revenue_USD", "sum"),
        )
        .sort_values(["Year", "Week"])
    )

    for row in range(8, 121):
        for col in range(1, 16):
            weekly_sheet.cell(row=row, column=col).value = None

    for idx, row_data in enumerate(weekly_agg.itertuples(index=False), start=8):
        if idx > 150:
            break
        weekly_sheet.cell(row=idx, column=1, value=idx - 7)
        weekly_sheet.cell(row=idx, column=2, value=row_data.Week_Start)
        weekly_sheet.cell(row=idx, column=3, value=row_data.Week_End)
        weekly_sheet.cell(row=idx, column=4, value="All Platforms")
        weekly_sheet.cell(row=idx, column=5, value=row_data.Impressions)
        weekly_sheet.cell(row=idx, column=6, value=row_data.Clicks)
        weekly_sheet.cell(row=idx, column=7, value=f"=IF(E{idx}=0,0,F{idx}/E{idx})")
        weekly_sheet.cell(row=idx, column=8, value=row_data.Conversions)
        weekly_sheet.cell(row=idx, column=9, value=f"=IF(F{idx}=0,0,H{idx}/F{idx})")
        weekly_sheet.cell(row=idx, column=10, value=row_data.Spend)
        weekly_sheet.cell(row=idx, column=11, value=f"=IF(E{idx}=0,0,(J{idx}/E{idx})*1000)")
        weekly_sheet.cell(row=idx, column=12, value=f"=IF(F{idx}=0,0,J{idx}/F{idx})")
        weekly_sheet.cell(row=idx, column=13, value=f"=IF(H{idx}=0,0,J{idx}/H{idx})")
        weekly_sheet.cell(row=idx, column=14, value=f"=IF(J{idx}=0,0,O{idx}/J{idx})")
        weekly_sheet.cell(row=idx, column=15, value=row_data.Revenue)

    last_week_row = min(7 + len(weekly_agg), 150)
    totals_row_w = last_week_row + 2
    weekly_sheet.cell(row=totals_row_w, column=1, value="TOTALS")
    weekly_sheet.cell(row=totals_row_w, column=5, value=f"=SUM(E8:E{last_week_row})")
    weekly_sheet.cell(row=totals_row_w, column=6, value=f"=SUM(F8:F{last_week_row})")
    weekly_sheet.cell(row=totals_row_w, column=8, value=f"=SUM(H8:H{last_week_row})")
    weekly_sheet.cell(row=totals_row_w, column=10, value=f"=SUM(J8:J{last_week_row})")
    weekly_sheet.cell(row=totals_row_w, column=15, value=f"=SUM(O8:O{last_week_row})")
    weekly_sheet.cell(row=totals_row_w, column=7, value=f"=IF(E{totals_row_w}=0,0,F{totals_row_w}/E{totals_row_w})")
    weekly_sheet.cell(row=totals_row_w, column=9, value=f"=IF(F{totals_row_w}=0,0,H{totals_row_w}/F{totals_row_w})")
    weekly_sheet.cell(row=totals_row_w, column=11, value=f"=IF(E{totals_row_w}=0,0,(J{totals_row_w}/E{totals_row_w})*1000)")
    weekly_sheet.cell(row=totals_row_w, column=12, value=f"=IF(F{totals_row_w}=0,0,J{totals_row_w}/F{totals_row_w})")
    weekly_sheet.cell(row=totals_row_w, column=13, value=f"=IF(H{totals_row_w}=0,0,J{totals_row_w}/H{totals_row_w})")
    weekly_sheet.cell(row=totals_row_w, column=14, value=f"=IF(J{totals_row_w}=0,0,O{totals_row_w}/J{totals_row_w})")

    output_path = Path(output_path) if output_path else excel_path
    wb.save(output_path)

    summary = {
        "records_processed": len(df),
        "daily_records": len(daily_agg),
        "weekly_records": len(weekly_agg),
        "start_date": start_date,
        "end_date": end_date,
        "total_spend": total_spend,
        "total_revenue": total_revenue,
        "platforms": platforms,
        "output_file": str(output_path),
    }

    roas = 0.0 if total_spend == 0 else total_revenue / total_spend
    print("=" * 60)
    print("UPDATE SUMMARY")
    print("=" * 60)
    print(f"CSV Records Processed: {summary['records_processed']}")
    print(f"Daily Records Created: {summary['daily_records']}")
    print(f"Weekly Records Created: {summary['weekly_records']}")
    print(f"Date Range: {start_date:%Y-%m-%d} to {end_date:%Y-%m-%d}")
    print(f"Platforms: {platform_str}")
    print(f"Total Spend: ${total_spend:,.2f}")
    print(f"Total Revenue: ${total_revenue:,.2f}")
    print(f"Overall ROAS: {roas:.2f}x")
    print("Output File:", output_path)
    print("=" * 60)

    return summary


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("csv", help="Path to normalized multi-platform CSV")
    parser.add_argument("template", help="Path to pacing tracker Excel template")
    parser.add_argument(
        "--output",
        "-o",
        help="Optional path for updated workbook (defaults to overwriting template)",
    )
    args = parser.parse_args(argv)

    try:
        update_pacing_tracker(args.csv, args.template, args.output)
    except Exception as exc:  # pragma: no cover
        print(f"Error updating tracker: {exc}", file=sys.stderr)
        raise
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
