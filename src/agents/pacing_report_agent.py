"""
Pacing Report Agent for generating Excel-based pacing reports.

This agent handles:
- Excel template validation and processing
- Daily/Weekly/Monthly aggregation of campaign data
- Automated report generation and scheduling
- Integration with existing campaign data sources
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List, Literal
from enum import Enum
from functools import lru_cache
import os

import pandas as pd
from openpyxl import load_workbook
from loguru import logger

from ..database.duckdb_manager import get_duckdb_manager
from .adaptive_sheet_populator import AdaptiveSheetPopulator


class AggregationLevel(str, Enum):
    """Supported aggregation levels for pacing reports."""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"


class PacingReportAgent:
    """Agent for generating Excel-based pacing reports with flexible aggregation."""
    
    def __init__(self, storage_dir: Path = None):
        """
        Initialize Pacing Report Agent.
        
        Args:
            storage_dir: Directory to store templates and generated reports
        """
        self.storage_dir = storage_dir or Path("reports/pacing")
        self.templates_dir = self.storage_dir / "templates"
        self.outputs_dir = self.storage_dir / "outputs"
        
        # Create directories
        self.templates_dir.mkdir(parents=True, exist_ok=True)
        self.outputs_dir.mkdir(parents=True, exist_ok=True)
        self.populator = AdaptiveSheetPopulator()
        self._jobs = {}  # Track background generation jobs
        self._template_cache = {}  # Cache for template validation
        
        logger.info(f"Initialized PacingReportAgent with storage: {self.storage_dir}")
    
    # ============ JOB MANAGEMENT (Progress Tracking) ============
    
    def get_job_status(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Get the status of a background generation job."""
        return self._jobs.get(job_id)
    
    def list_all_jobs(self, limit: int = 20) -> List[Dict[str, Any]]:
        """
        List all jobs with their status for progress monitoring.
        
        Returns: List of job summaries sorted by most recent
        """
        jobs = []
        for job_id, job_data in self._jobs.items():
            jobs.append({
                "job_id": job_id,
                **job_data
            })
        # Sort by updated_at descending
        jobs.sort(key=lambda x: x.get('updated_at', ''), reverse=True)
        return jobs[:limit]
    
    def cancel_job(self, job_id: str) -> bool:
        """Request cancellation of a running job."""
        if job_id in self._jobs:
            self._jobs[job_id]['status'] = 'cancelled'
            self._jobs[job_id]['message'] = 'Job cancelled by user'
            return True
        return False
    
    def cleanup_old_jobs(self, max_age_hours: int = 24):
        """Remove old completed jobs from memory."""
        cutoff = datetime.now()
        to_remove = []
        for job_id, job_data in self._jobs.items():
            if job_data.get('status') in ['completed', 'failed', 'cancelled']:
                try:
                    updated = datetime.fromisoformat(job_data.get('updated_at', ''))
                    if (cutoff - updated).total_seconds() > max_age_hours * 3600:
                        to_remove.append(job_id)
                except:
                    pass
        for job_id in to_remove:
            del self._jobs[job_id]
        if to_remove:
            logger.info(f"Cleaned up {len(to_remove)} old jobs")

    def _update_job_status(self, job_id: str, status: str, progress: float = 0, message: str = "", result: Any = None):
        """Update the status of a background job."""
        if job_id not in self._jobs:
            self._jobs[job_id] = {"created_at": datetime.now().isoformat()}
        
        self._jobs[job_id].update({
            "status": status,
            "progress": progress,
            "message": message,
            "updated_at": datetime.now().isoformat()
        })
        if result:
            self._jobs[job_id]["result"] = result
        
        logger.debug(f"Job {job_id} updated: {status} ({progress}%) - {message}")
    
    # ============ TEMPLATE CACHING ============
    
    def _get_template_mtime(self, template_path: Path) -> float:
        """Get file modification time for cache invalidation."""
        try:
            return os.path.getmtime(template_path)
        except:
            return 0.0
    
    def clear_template_cache(self):
        """Clear all cached template validations."""
        self._template_cache.clear()
        logger.info("Template validation cache cleared")
    
    def validate_template(self, template_path: Path, use_cache: bool = True) -> Dict[str, Any]:
        """
        Intelligently validate Excel template structure with fuzzy matching.
        Uses caching based on file modification time for faster repeated validations.
        
        Args:
            template_path: Path to Excel template file
            use_cache: Whether to use cached result (default: True)
            
        Returns:
            Validation result with status, details, and suggestions
        """
        path_str = str(template_path)
        mtime = self._get_template_mtime(template_path)
        
        # Check cache
        if use_cache and path_str in self._template_cache:
            cached_mtime, cached_result = self._template_cache[path_str]
            if cached_mtime == mtime:
                logger.debug(f"Template validation cache hit: {template_path.name}")
                return cached_result
        try:
            wb = load_workbook(template_path)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                return {
                    "valid": False,
                    "error": "Excel file contains no sheets"
                }
            
            # Recommended sheet patterns (case-insensitive, fuzzy matching)
            recommended_patterns = {
                "daily": ["daily pacing", "daily", "day", "daily data", "daily report"],
                "weekly": ["weekly pacing", "weekly", "week", "weekly data", "weekly report"],
            }
            
            # Smart sheet detection
            detected_sheets = {
                "daily": None,
                "weekly": None,
                "other": []
            }
            
            for sheet in sheet_names:
                sheet_lower = sheet.lower().strip()
                
                # Check for daily patterns
                if any(pattern in sheet_lower for pattern in recommended_patterns["daily"]):
                    detected_sheets["daily"] = sheet
                # Check for weekly patterns
                elif any(pattern in sheet_lower for pattern in recommended_patterns["weekly"]):
                    detected_sheets["weekly"] = sheet
                else:
                    detected_sheets["other"].append(sheet)
            
            # Build validation response
            warnings = []
            suggestions = []
            
            if not detected_sheets["daily"]:
                warnings.append("No 'Daily Pacing' sheet detected")
                suggestions.append("Consider adding a sheet named 'Daily Pacing' for daily aggregated data")
            
            if not detected_sheets["weekly"]:
                warnings.append("No 'Weekly Pacing' sheet detected")
                suggestions.append("Consider adding a sheet named 'Weekly Pacing' for weekly aggregated data")
            
            # Analyze sheet structure
            sheet_info = []
            for sheet_name in sheet_names:
                try:
                    sheet = wb[sheet_name]
                    row_count = sheet.max_row
                    col_count = sheet.max_column
                    sheet_info.append({
                        "name": sheet_name,
                        "rows": row_count,
                        "columns": col_count,
                        "type": "daily" if sheet_name == detected_sheets["daily"] 
                               else "weekly" if sheet_name == detected_sheets["weekly"]
                               else "other"
                    })
                except Exception as e:
                    logger.warning(f"Could not analyze sheet {sheet_name}: {e}")
            
            # Determine validation status
            has_pacing_sheets = detected_sheets["daily"] or detected_sheets["weekly"]
            
            result = {
                "valid": True,  # Always accept Excel files
                "sheets": sheet_names,
                "sheet_count": len(sheet_names),
                "detected_sheets": {
                    "daily": detected_sheets["daily"],
                    "weekly": detected_sheets["weekly"],
                    "other": detected_sheets["other"]
                },
                "sheet_info": sheet_info,
                "message": "Template uploaded successfully"
            }
            
            if warnings:
                result["warnings"] = warnings
            
            if suggestions:
                result["suggestions"] = suggestions
            
            if has_pacing_sheets:
                result["message"] = f"Template is ready! Detected {len([s for s in [detected_sheets['daily'], detected_sheets['weekly']] if s])} pacing sheet(s)"
            else:
                result["message"] = "Template uploaded. Note: No standard pacing sheets detected, but you can still use this template"
            
            logger.info(f"Template validated: {len(sheet_names)} sheets, daily={detected_sheets['daily']}, weekly={detected_sheets['weekly']}")
            
            # Store in cache
            self._template_cache[path_str] = (mtime, result)
            
            return result
            
        except Exception as e:
            logger.error(f"Template validation failed: {e}")
            return {
                "valid": False,
                "error": f"Invalid Excel file: {str(e)}. Please ensure the file is a valid .xlsx Excel file."
            }
    
    def save_template(self, file_content: bytes, filename: str) -> Path:
        """
        Save uploaded template file.
        
        Args:
            file_content: Template file content
            filename: Original filename
            
        Returns:
            Path to saved template
        """
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = f"template_{timestamp}_{filename}"
        template_path = self.templates_dir / safe_filename
        
        # Save file
        template_path.write_bytes(file_content)
        logger.info(f"Saved template: {template_path}")
        
        return template_path
    
    def fetch_campaign_data(
        self,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        filters: Optional[Dict[str, Any]] = None
    ) -> pd.DataFrame:
        """
        Fetch campaign data from database.
        
        Args:
            start_date: Start_date (YYYY-MM-DD)
            end_date: End date (YYYY-MM-DD)
            filters: Additional filters (platform, campaign_id, etc.)
            
        Returns:
            DataFrame with campaign data
        """
        db = get_duckdb_manager()
        
        # Build filters dict for DuckDBManager
        db_filters = filters.copy() if filters else {}
        
        # Add date filters if provided
        if start_date:
            db_filters['start_date'] = start_date
        if end_date:
            db_filters['end_date'] = end_date
        
        # Fetch data using DuckDBManager's get_campaigns method
        df = db.get_campaigns(filters=db_filters, limit=1000000)
        
        if df.empty:
            logger.warning("No campaign data found for specified criteria")
        else:
            logger.info(f"Fetched {len(df)} campaign records")
        
        return df
    
    def normalize_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        INTELLIGENT DATA NORMALIZER
        
        Uses semantic understanding and pattern matching to automatically 
        discover and map columns - no hardcoding required.
        
        Features:
        - Fuzzy column name matching
        - Year-suffix column detection and date-aware merging
        - Value-based type inference
        - Self-documenting mapping decisions
        
        Expected output: Date, Platform, Campaign, Impressions, Clicks, Spend_USD, Conversions, Revenue_USD
        """
        logger.info("🧠 Intelligent Data Normalizer starting...")
        
        df_normalized = df.copy()
        original_columns = list(df.columns)
        df_normalized.columns = df_normalized.columns.str.lower().str.strip()
        
        # Define semantic patterns for each target field
        # Each pattern list is ordered by priority (most specific first)
        semantic_patterns = {
            'Date': ['date', 'day', 'timestamp', 'time', 'period', 'fecha'],
            'Platform': ['platform', 'channel', 'source', 'network', 'publisher', 'medio'],
            'Campaign': ['campaign', 'campaign_name', 'ad_group', 'adgroup', 'initiative'],
            'Impressions': ['impression', 'impress', 'views', 'reach', 'eyeballs'],
            'Clicks': ['click', 'clic', 'visits', 'sessions'],
            'Spend_USD': ['spend', 'cost', 'spent', 'budget', 'investment', 'gasto', 'amount'],
            'Conversions': ['conversion', 'convert', 'lead', 'action', 'goal', 'site visit', 'purchase', 'signup'],
            'Revenue_USD': ['revenue', 'income', 'sales', 'value', 'earnings', 'ingreso']
        }
        
        mapping_log = []
        mapped_targets = {}
        
        # PHASE 1: Detect year-suffixed columns (e.g., revenue_2024, revenue_2025)
        year_pattern = r'(.+?)_?(20\d{2})$'
        year_columns = {}
        import re
        
        for col in df_normalized.columns:
            match = re.match(year_pattern, col.replace(' ', '_'))
            if match:
                base_name = match.group(1).replace('_', ' ').strip()
                year = int(match.group(2))
                if base_name not in year_columns:
                    year_columns[base_name] = {}
                year_columns[base_name][year] = col
        
        if year_columns:
            logger.info(f"📅 Detected year-suffixed columns: {list(year_columns.keys())}")
        
        # PHASE 2: Parse date column first (needed for year-aware merging)
        # ROBUST DATE PARSING - handles multiple formats
        date_col = None
        for col in df_normalized.columns:
            for pattern in semantic_patterns['Date']:
                if pattern in col and 'day of' not in col:  # Exclude "day of week"
                    date_col = col
                    break
            if date_col:
                break
        
        if date_col:
            # Multi-format date parsing for robustness
            raw_dates = df_normalized[date_col]
            parsed_dates = None
            successful_format = None
            
            # Common date formats to try (ordered by likelihood)
            date_formats = [
                '%d/%m/%y',      # DD/MM/YY (e.g., 01/01/24)
                '%m/%d/%y',      # MM/DD/YY (e.g., 01/31/24) 
                '%d/%m/%Y',      # DD/MM/YYYY (e.g., 01/01/2024)
                '%m/%d/%Y',      # MM/DD/YYYY (e.g., 01/31/2024)
                '%Y-%m-%d',      # ISO format (e.g., 2024-01-01)
                '%Y/%m/%d',      # YYYY/MM/DD
                '%d-%m-%Y',      # DD-MM-YYYY
                '%d-%m-%y',      # DD-MM-YY
                '%Y%m%d',        # YYYYMMDD compact
            ]
            
            # Try each format until one works for at least 90% of non-null values
            for fmt in date_formats:
                try:
                    test_parsed = pd.to_datetime(raw_dates, format=fmt, errors='coerce')
                    valid_count = test_parsed.notna().sum()
                    total_count = raw_dates.notna().sum()
                    
                    if total_count > 0 and valid_count / total_count >= 0.9:
                        # Validate parsed dates are in reasonable range (2000-2050)
                        years = test_parsed.dt.year.dropna()
                        if len(years) > 0 and years.min() >= 2000 and years.max() <= 2050:
                            parsed_dates = test_parsed
                            successful_format = fmt
                            logger.info(f"📅 Date format detected: '{fmt}' ({valid_count}/{total_count} valid)")
                            break
                except Exception:
                    continue
            
            # Fallback: let pandas infer (dayfirst=True for DD/MM format priority)
            if parsed_dates is None:
                logger.warning(f"⚠️ No explicit date format matched, attempting auto-inference...")
                try:
                    # Try dayfirst=True first (more common internationally)
                    parsed_dates = pd.to_datetime(raw_dates, dayfirst=True, errors='coerce')
                    years = parsed_dates.dt.year.dropna()
                    
                    # Validate the inferred dates
                    if len(years) == 0 or years.min() < 2000 or years.max() > 2050:
                        # Try dayfirst=False (US format)
                        parsed_dates = pd.to_datetime(raw_dates, dayfirst=False, errors='coerce')
                        successful_format = 'auto-infer (US format)'
                    else:
                        successful_format = 'auto-infer (dayfirst)'
                except Exception as e:
                    logger.error(f"❌ Date parsing failed completely: {e}")
                    parsed_dates = pd.NaT
            
            df_normalized['_temp_date'] = parsed_dates
            df_normalized['_year'] = df_normalized['_temp_date'].dt.year
            
            # Also store properly parsed Date for output
            df_normalized['Date'] = parsed_dates
            
            # Log parsing stats
            valid_dates = parsed_dates.notna().sum() if parsed_dates is not None else 0
            total_dates = len(raw_dates)
            mapping_log.append(f"✓ Date: '{date_col}' → Date (format: {successful_format}, {valid_dates}/{total_dates} parsed)")
            mapped_targets['Date'] = date_col
        
        # PHASE 3: Map direct columns using semantic matching (BEFORE year-suffix merge)
        # This ensures we use 'impressions' directly instead of 'reach_2024/2025'
        for target_field, patterns in semantic_patterns.items():
            if target_field in mapped_targets:
                continue
            
            best_match = None
            best_score = 0
            
            for col in df_normalized.columns:
                if col.startswith('_'):  # Skip temp columns
                    continue
                # Skip year-suffixed columns for now
                if re.match(r'.+_20\d{2}$', col.replace(' ', '_')):
                    continue
                    
                # Calculate match score
                for i, pattern in enumerate(patterns):
                    if pattern in col:
                        score = len(patterns) - i  # Higher priority patterns get higher scores
                        # Boost exact matches
                        if col == pattern or col.replace('_', ' ') == pattern:
                            score += 10
                        if score > best_score:
                            best_score = score
                            best_match = col
            
            if best_match:
                if target_field not in df_normalized.columns:
                    df_normalized[target_field] = df_normalized[best_match]
                mapped_targets[target_field] = best_match
                mapping_log.append(f"✓ {target_field}: '{best_match}' (score: {best_score})")
        
        # PHASE 4: Merge year-suffixed columns for REMAINING unmapped fields
        for base_name, year_cols in year_columns.items():
            # Determine target field
            target_field = None
            for target, patterns in semantic_patterns.items():
                if any(p in base_name.lower() for p in patterns):
                    target_field = target
                    break
            
            # Only use year-suffix merge if not already mapped
            if target_field and target_field not in mapped_targets:
                # Create merged column based on date year
                df_normalized[target_field] = 0.0
                years_mapped = []
                
                for year, col_name in sorted(year_cols.items()):
                    mask = df_normalized['_year'] == year
                    df_normalized.loc[mask, target_field] = pd.to_numeric(
                        df_normalized.loc[mask, col_name], errors='coerce'
                    ).fillna(0)
                    years_mapped.append(f"{year}:{mask.sum()} rows")
                
                mapped_targets[target_field] = f"merged({list(year_cols.values())})"
                mapping_log.append(f"✓ {target_field}: Date-aware merge from {list(year_cols.values())} ({', '.join(years_mapped)})")
        
        # PHASE 5: Value-based inference for unmapped fields
        required_cols = ['Date', 'Platform', 'Campaign', 'Impressions', 'Clicks', 'Spend_USD', 'Conversions', 'Revenue_USD']
        missing = [c for c in required_cols if c not in mapped_targets]
        
        if missing:
            logger.info(f"🔍 Attempting value-based inference for: {missing}")
            
            for col in df_normalized.columns:
                if col.startswith('_') or col in [mapped_targets.get(t) for t in mapped_targets]:
                    continue
                
                sample = df_normalized[col].dropna().head(100)
                if len(sample) == 0:
                    continue
                
                # Infer type from values
                if 'Platform' in missing and sample.dtype == 'object':
                    unique_ratio = sample.nunique() / len(sample)
                    if 0.01 < unique_ratio < 0.5:  # Categorical with moderate cardinality
                        df_normalized['Platform'] = df_normalized[col]
                        mapped_targets['Platform'] = col
                        mapping_log.append(f"✓ Platform: '{col}' (inferred: categorical)")
                        missing.remove('Platform')
        
        # PHASE 6: Fill remaining missing columns with defaults
        for col in required_cols:
            if col not in df_normalized.columns:
                if col in ['Platform', 'Campaign']:
                    df_normalized[col] = 'Unknown'
                    mapping_log.append(f"⚠ {col}: Default 'Unknown'")
                else:
                    df_normalized[col] = 0
                    mapping_log.append(f"⚠ {col}: Default 0")
        
        # PHASE 7: Type coercion
        df_normalized['Date'] = pd.to_datetime(df_normalized['Date'], errors='coerce')
        df_normalized['Platform'] = df_normalized['Platform'].astype(str)
        df_normalized['Campaign'] = df_normalized['Campaign'].astype(str)
        
        numeric_cols = ['Impressions', 'Clicks', 'Spend_USD', 'Conversions', 'Revenue_USD']
        for col in numeric_cols:
            df_normalized[col] = pd.to_numeric(df_normalized[col], errors='coerce').fillna(0)
        
        # Clean up temp columns
        df_normalized = df_normalized.drop(columns=['_temp_date', '_year'], errors='ignore')
        
        # Log complete mapping report
        logger.info("=" * 60)
        logger.info("🧠 INTELLIGENT NORMALIZER - MAPPING REPORT")
        logger.info("=" * 60)
        logger.info(f"📥 Input: {len(original_columns)} columns")
        for log_entry in mapping_log:
            logger.info(f"   {log_entry}")
        
        # Identify "Custom Dimensions" (original columns that weren't mapped to standard ones)
        mapped_source_cols = set()
        for src in mapped_targets.values():
            if isinstance(src, str):
                if src.startswith('merged('):
                    # Extract column names from merged(...)
                    inner = src[7:-1]
                    for part in inner.split(','):
                        mapped_source_cols.add(part.strip().strip("'[]"))
                else:
                    mapped_source_cols.add(src)
        
        custom_dimensions = [c for c in original_columns if c.lower().strip() not in mapped_source_cols]
        if custom_dimensions:
            logger.info(f"✨ Preserved {len(custom_dimensions)} custom dimensions: {custom_dimensions}")
        
        logger.info(f"📤 Output: {len(df_normalized.columns)} columns (all original + normalized)")
        logger.info("=" * 60)
        
        return df_normalized
    
    def aggregate_data(
        self,
        df: pd.DataFrame,
        level: AggregationLevel,
        dimensions: List[str] = None
    ) -> pd.DataFrame:
        """
        Aggregate data by specified level and dimensions.
        
        Args:
            df: Normalized campaign data
            level: Aggregation level (daily, weekly, monthly)
            dimensions: List of dimensions to group by (default: Platform, Campaign)
            
        Returns:
            Aggregated DataFrame
        """
        df = df.copy()
        
        # Determine base time grouping
        if level == AggregationLevel.DAILY:
            time_cols = ['Date']
            # Default dimensions for Daily: Platform, Campaign
            if dimensions is None:
                dimensions = ['Platform', 'Campaign']
        elif level == AggregationLevel.WEEKLY:
            df['Week'] = df['Date'].dt.isocalendar().week
            df['Year'] = df['Date'].dt.isocalendar().year
            df['Month'] = df['Date'].dt.month
            time_cols = ['Year', 'Week', 'Month']
            # Default dimensions for Weekly: Platform
            if dimensions is None:
                dimensions = ['Platform']
        elif level == AggregationLevel.MONTHLY:
            df['Month'] = df['Date'].dt.to_period('M')
            time_cols = ['Month']
            # Default dimensions for Monthly: None (Budget Pacing Dashboard wants totals)
            if dimensions is None:
                dimensions = []
        else:
            raise ValueError(f"Unsupported aggregation level: {level}")

        # Ensure all dimensions exist in DataFrame (fallback to 'Unknown' if not)
        safe_dimensions = []
        for d in dimensions:
            if d in df.columns:
                safe_dimensions.append(d)
                # Ensure it's a string for grouping if it's a dimension
                if df[d].dtype == 'object' or df[d].dtype.name == 'category':
                    df[d] = df[d].astype(str)
            else:
                logger.warning(f"⚠️ Dimension '{d}' requested but not found in data. Using 'Unknown'.")
                df[d] = 'Unknown'
                safe_dimensions.append(d)

        # Merge time and dimensions for final grouping
        group_cols = time_cols + safe_dimensions
        
        # Aggregate metrics
        metric_agg = {
            'Impressions': 'sum',
            'Clicks': 'sum',
            'Spend_USD': 'sum',
            'Conversions': 'sum',
            'Revenue_USD': 'sum'
        }
        
        # Add any other numeric columns that might be present in the original data and not already captured
        # (This supports custom metrics if provided in certain templates)
        # For now we stick to the core ones to keep it clean
        
        agg_df = df.groupby(group_cols, as_index=False).agg(metric_agg)
        
        # Post-aggregation processing
        if level == AggregationLevel.DAILY:
            # Add date derivatives
            agg_df['Year'] = agg_df['Date'].dt.year
            agg_df['Month'] = agg_df['Date'].dt.month
            agg_df['Week'] = agg_df['Date'].dt.isocalendar().week
            agg_df['Day of Week'] = agg_df['Date'].dt.day_name()
            # Format Date as YYYY-MM-DD string
            agg_df['Date'] = agg_df['Date'].dt.strftime('%Y-%m-%d')
            agg_df = agg_df.sort_values(group_cols)
            
        elif level == AggregationLevel.WEEKLY:
            # Add range dates for weekly display
            weekly_dates = df.groupby(['Year', 'Week']).agg({'Date': ['min', 'max']}).reset_index()
            weekly_dates.columns = ['Year', 'Week', 'Date_Start', 'Date_End']
            agg_df = agg_df.merge(weekly_dates, on=['Year', 'Week'])
            # Compatibility: use 'Date' for start date
            agg_df = agg_df.rename(columns={'Date_Start': 'Date', 'Date_End': 'Week_End'})
            agg_df['Date'] = pd.to_datetime(agg_df['Date']).dt.strftime('%Y-%m-%d')
            agg_df['Week_End'] = pd.to_datetime(agg_df['Week_End']).dt.strftime('%Y-%m-%d')
            agg_df = agg_df.sort_values(group_cols)
            
        elif level == AggregationLevel.MONTHLY:
            # Create Year-Month column correctly
            agg_df['Year-Month'] = agg_df['Month'].astype(str)
            agg_df = agg_df.rename(columns={'Month': 'Date'})
            agg_df['Date'] = agg_df['Date'].astype(str)
            
            # BUDGET PACING CALCULATIONS (specific to monthly totals)
            if not dimensions or 'Platform' not in dimensions:
                MONTHLY_BUDGET = 100000
                agg_df['Monthly Spend'] = agg_df['Spend_USD']
                agg_df['Budget'] = MONTHLY_BUDGET
                agg_df['Variance'] = agg_df['Budget'] - agg_df['Spend_USD']
                agg_df['Variance %'] = ((agg_df['Variance'] / agg_df['Budget']) * 100).round(2)
                
                def get_pacing_status(variance_pct):
                    if abs(variance_pct) <= 5: return 'On Track'
                    elif variance_pct > 5: return 'Under Pacing'
                    else: return 'Over Pacing'
                
                def get_alert(variance_pct):
                    abs_var = abs(variance_pct)
                    if abs_var <= 10: return '🟢'
                    elif abs_var <= 25: return '🟡'
                    else: return '🔴'
                
                agg_df['Status'] = agg_df['Variance %'].apply(get_pacing_status)
                agg_df['Alert'] = agg_df['Variance %'].apply(get_alert)

        # Common Performance Metrics
        agg_df['CTR'] = (agg_df['Clicks'] / agg_df['Impressions']).replace([float('inf'), -float('inf')], 0).fillna(0)
        agg_df['CPC'] = (agg_df['Spend_USD'] / agg_df['Clicks']).replace([float('inf'), -float('inf')], 0).fillna(0)
        agg_df['CPM'] = (agg_df['Spend_USD'] / agg_df['Impressions'] * 1000).replace([float('inf'), -float('inf')], 0).fillna(0)
        agg_df['CPA'] = (agg_df['Spend_USD'] / agg_df['Conversions']).replace([float('inf'), -float('inf')], 0).fillna(0)
        agg_df['CVR'] = (agg_df['Conversions'] / agg_df['Clicks']).replace([float('inf'), -float('inf')], 0).fillna(0)
        agg_df['ROAS'] = (agg_df['Revenue_USD'] / agg_df['Spend_USD']).replace([float('inf'), -float('inf')], 0).fillna(0)
        
        # Finance metrics
        agg_df['Profit'] = agg_df['Revenue_USD'] - agg_df['Spend_USD']
        agg_df['Margin %'] = ((agg_df['Profit'] / agg_df['Revenue_USD']) * 100).replace([float('inf'), -float('inf')], 0).fillna(0).round(2)
        
        # Aliases
        agg_df['Revenue'] = agg_df['Revenue_USD']
        
        logger.info(f"📊 {level.value} aggregation complete: {len(agg_df)} records, dimensions: {dimensions}")
        return agg_df
    
    def _safe_number(self, value):
        """Convert value to float safely."""
        try:
            if value is None: return 0.0
            if isinstance(value, (int, float)): return float(value)
            # Remove currency symbols and commas
            clean_val = str(value).replace('$', '').replace(',', '').strip()
            return float(clean_val)
        except:
            return 0.0

    def _get_best_data_for_table(
        self, 
        headers: List[str], 
        source_data: pd.DataFrame,
        sheet_type_hint: str = None
    ) -> pd.DataFrame:
        """
        UNIVERSAL DIMENSION SCAVENGER
        
        Dynamically detects which dimensions are needed based on table headers
        and produces the optimal aggregation.
        """
        logger.info(f"🔍 Scavenging dimensions from headers: {headers}")
        
        # 1. Identify Time Grain
        grain = AggregationLevel.DAILY
        hdr_lower = [h.lower() for h in headers]
        
        if any(h in hdr_lower for h in ['month', 'year-month', 'period']):
            grain = AggregationLevel.MONTHLY
        elif any(h in hdr_lower for h in ['week', 'iso_week', 'week_end']):
            grain = AggregationLevel.WEEKLY
        elif sheet_type_hint == 'monthly':
            grain = AggregationLevel.MONTHLY
        elif sheet_type_hint == 'weekly':
            grain = AggregationLevel.WEEKLY
            
        logger.info(f"   → Detected grain: {grain.value}")
        
        # 2. Identify Dimensions (Anything that isn't a date or a metric)
        metrics = [
            'impressions', 'clicks', 'spend', 'cost', 'spend_usd', 'conversions', 
            'revenue', 'revenue_usd', 'ctr', 'cpc', 'cpm', 'cpa', 'cvr', 'roas',
            'profit', 'margin', 'margin %', 'budget', 'variance', 'status', 'alert'
        ]
        time_terms = ['date', 'day', 'month', 'week', 'year', 'period', 'year-month', 'day of week']
        
        needed_dimensions = []
        for h in headers:
            h_clean = h.lower().strip()
            if h_clean in time_terms or h_clean in metrics:
                continue
            
            # This is likely a dimension. Check if it exists in source data.
            # We check both exact match and fuzzy match
            match = None
            for col in source_data.columns:
                if col.lower().strip() == h_clean:
                    match = col
                    break
            
            if match:
                needed_dimensions.append(match)
                logger.info(f"   ✨ Dynamic Dimension Found: '{match}' matches header '{h}'")
            else:
                # Fallback: check normalized names
                if h_clean in ['platform', 'channel', 'source']:
                    needed_dimensions.append('Platform')
                elif h_clean in ['campaign', 'ad group', 'initiative']:
                    needed_dimensions.append('Campaign')
        
        # Remove duplicates
        needed_dimensions = list(dict.fromkeys(needed_dimensions))
        
        # 3. Aggregate data
        return self.aggregate_data(source_data, grain, dimensions=needed_dimensions)
        
    def populate_daily_sheet(
        self,
        sheet,
        daily_data: pd.DataFrame,
        campaign_info: Dict[str, Any]
    ):
        """
        Intelligently populate sheet with daily data using adaptive detection.
        
        This method automatically detects ALL tables in the template and writes data to each.
        """
        try:
            logger.info(f"Populating sheet '{sheet.title}' with {len(daily_data)} daily records")
            
            # Use adaptive populator with multi-table detection
            populator = AdaptiveSheetPopulator()
            
            # First, try to detect and populate all tables
            result = populator.populate_all_tables(sheet, daily_data, clear_existing=True)
            
            if result.get('tables_populated', 0) > 0:
                logger.success(f"Populated {result.get('tables_populated', 0)} tables with {result.get('total_rows_written', 0)} total rows in '{sheet.title}'")
            elif result.get('success'):
                logger.success(f"Successfully populated {result.get('rows_written', 0)} rows in '{sheet.title}'")
            else:
                logger.error(f"Failed to populate sheet: {result.get('error')}")
            
            return result
            
        except Exception as e:
            logger.exception(f"Error populating daily sheet: {e}")
            return {"success": False, "error": str(e)}
    
    def populate_weekly_sheet(
        self,
        sheet,
        weekly_data: pd.DataFrame
    ):
        """
        Intelligently populate sheet with weekly data using adaptive detection.
        Detects ALL tables in the sheet and populates each one.
        """
        try:
            logger.info(f"Populating sheet '{sheet.title}' with {len(weekly_data)} weekly records")
            
            # Use adaptive populator with multi-table detection
            result = self.populator.populate_all_tables(sheet, weekly_data, clear_existing=True)
            
            if result.get('tables_populated', 0) > 0:
                logger.success(f"Populated {result.get('tables_populated', 0)} tables in '{sheet.title}'")
            elif result.get('success'):
                logger.success(f"Successfully populated {result.get('rows_written', 0)} rows in '{sheet.title}'")
            else:
                logger.error(f"Failed to populate weekly sheet: {result.get('error')}")
            
            return result
            
        except Exception as e:
            logger.exception(f"Error populating weekly sheet: {e}")
            return {"success": False, "error": str(e)}
    
    def generate_report(
        self,
        template_path: Path,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        aggregation: AggregationLevel = AggregationLevel.DAILY,
        filters: Optional[Dict[str, Any]] = None,
        output_filename: Optional[str] = None,
        job_id: Optional[str] = None,
        max_daily_rows: Optional[int] = None  # None = no limit, set to number to limit Campaign Data rows
    ) -> Dict[str, Any]:
        """
        Generate pacing report from template and data.
        
        Args:
            template_path: Path to Excel template
            start_date: Start date for data (YYYY-MM-DD)
            end_date: End date for data (YYYY-MM-DD)
            aggregation: Aggregation level
            filters: Additional data filters
            output_filename: Custom output filename
            job_id: Optional ID for background task tracking
            max_daily_rows: Maximum rows for Campaign Data sheet (default 10000 for fast generation)
            
        Returns:
            Report generation result with file path and summary
        """
        if job_id:
            self._update_job_status(job_id, "processing", 5, "Fetching campaign data...")
            
        try:
            logger.info(f"Generating {aggregation.value} pacing report (Job: {job_id})...")
            
            # Fetch and normalize data
            raw_data = self.fetch_campaign_data(start_date, end_date, filters)
            if raw_data.empty:
                return {
                    "success": False,
                    "error": "No data available for specified criteria"
                }
            
            normalized_data = self.normalize_data(raw_data)
            
            if job_id:
                self._update_job_status(job_id, "processing", 15, f"Data normalized ({len(raw_data)} records). Loading template...")

            
            # Force GC before loading large workbook
            import gc
            gc.collect()
            
            # Load template - use data_only=False to preserve formulas
            logger.info(f"Loading template: {template_path}")
            wb = load_workbook(template_path)
            gc.collect()
            
            # PERFORMANCE: Clean up data sheets if template has excessive pre-existing data
            # This handles templates that were previously generated and saved with data
            # Only cleanup if max_daily_rows is set (not None)
            if max_daily_rows is not None and max_daily_rows > 0:
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet_type = self._identify_sheet_type(sheet_name)
                    if sheet_type in ('daily', 'weekly', 'monthly') and sheet.max_row > max_daily_rows + 100:
                        logger.info(f"⚡ Cleaning template sheet '{sheet_name}': removing {sheet.max_row - 2} excess data rows")
                        # Delete all rows except header (row 1) - much faster than cell-by-cell clearing
                        sheet.delete_rows(2, sheet.max_row - 1)
                        logger.info(f"   → Sheet reduced to {sheet.max_row} rows")
            
            # Generate aggregations
            if job_id:
                self._update_job_status(job_id, "processing", 30, "Aggregating metrics...")
                
            daily_data = self.aggregate_data(normalized_data, AggregationLevel.DAILY)
            weekly_data = self.aggregate_data(normalized_data, AggregationLevel.WEEKLY)
            monthly_data = self.aggregate_data(normalized_data, AggregationLevel.MONTHLY)
            
            # Campaign info
            campaign_info = {
                'name': 'Multi-Platform Campaign Report',
                'platforms': ' / '.join(sorted(normalized_data['Platform'].unique())),
                'campaign_id': f"Multi_Campaign_{datetime.now():%Y%m%d_%H%M%S}",
                'start_date': normalized_data['Date'].min(),
                'end_date': normalized_data['Date'].max(),
                'total_spend': normalized_data['Spend_USD'].sum()
            }
            
            # Populate ALL sheets intelligently - SEQUENTIAL with QA logging
            if job_id:
                self._update_job_status(job_id, "processing", 50, "Writing data to Excel tabs...")
                
            sheets_populated = []
            sheets_failed = []
            qa_report = []
            
            logger.info(f"🔄 Processing {len(wb.sheetnames)} sheets sequentially...")
            
            for idx, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]
                sheet_type = self._identify_sheet_type(sheet_name)
                
                progress = 50 + int((idx / len(wb.sheetnames)) * 35)
                if job_id:
                    self._update_job_status(job_id, "processing", progress, f"Populating '{sheet_name}'...")
                
                try:
                    result = None
                    if sheet_type in ('daily', 'weekly', 'monthly'):
                        # UNIVERSAL DIMENSION DISCOVERY
                        # Scavenge dimensions and grain directly from the sheet headers
                        logger.info(f"🔍 Scavenging dimensions for '{sheet_name}' ({sheet_type})...")
                        
                        # Peek at headers to determine optimal data
                        headers = []
                        # Scan first 10 rows for anything looking like headers
                        for r in range(1, 11):
                            row_vals = [str(sheet.cell(r, c).value or "").strip() for c in range(1, 20)]
                            if any(v for v in row_vals):
                                headers.extend([v for v in row_vals if v])
                        
                        target_data = self._get_best_data_for_table(headers, normalized_data, sheet_type_hint=sheet_type)
                        
                        # Apply row limit for performance (keep most recent dates) - only if limit is set
                        if sheet_type == 'daily' and max_daily_rows is not None and max_daily_rows > 0 and len(target_data) > max_daily_rows:
                            logger.info(f"⚡ Limiting daily data from {len(target_data)} to {max_daily_rows} rows (most recent)")
                            target_data = target_data.sort_values(['Date', 'Platform', 'Campaign'], ascending=[False, True, True]).head(max_daily_rows)
                        
                        result = self.populator.populate_all_tables(sheet, target_data, clear_existing=True)
                        sheets_populated.append(f"{sheet_type}: '{sheet_name}'")
                    elif sheet_type == 'formula':
                        # Formula-based sheets (like Pivot Analysis) use SUMIF/SUMIFS to reference other sheets
                        # Update Pivot Analysis with actual channels from data before preserving formulas
                        update_result = self._update_pivot_analysis_channels(sheet, daily_data)
                        if update_result.get('success'):
                            channels_count = update_result.get('channels_updated', 0)
                            qa_report.append(f"📊 '{sheet_name}' updated with {channels_count} channels from data")
                            logger.info(f"📊 Updated '{sheet_name}' with {channels_count} channels, formulas will auto-calculate")
                        else:
                            qa_report.append(f"📊 '{sheet_name}' preserved (formula-based pivot sheet)")
                            logger.info(f"📊 Skipping '{sheet_name}' - formula-based sheet, formulas will auto-calculate")
                    else:
                        # Fallback: Check if it's the very first sheet and nothing else has been populated
                        if not sheets_populated and sheet_name == wb.sheetnames[0]:
                            result = self.populate_daily_sheet(sheet, daily_data, campaign_info)
                            sheets_populated.append(f"fallback(daily): '{sheet_name}'")
                        else:
                            qa_report.append(f"⏭️ Skipped '{sheet_name}' (no type match)")
                    
                    # Log QA result - enhanced to show tables count
                    if result:
                        if result.get('success'):
                            tables_count = result.get('tables_populated', 0)
                            rows = result.get('total_rows_written', result.get('rows_written', 0))
                            cols = result.get('columns_mapped', 0)
                            if tables_count > 0:
                                qa_report.append(f"✅ '{sheet_name}': {tables_count} tables, {rows} rows")
                            else:
                                qa_report.append(f"✅ '{sheet_name}': {rows} rows, {cols} cols")
                            # Capture detailed QA for learning
                            qa_details = result.get('qa_details', {})
                            if qa_details.get('unmapped_template_columns'):
                                qa_report.append(f"   ⚠️ Unmapped: {qa_details['unmapped_template_columns']}")
                            if qa_details.get('suggestions'):
                                for suggestion in qa_details['suggestions'][:2]:
                                    qa_report.append(f"   💡 {suggestion}")
                        else:
                            qa_report.append(f"❌ '{sheet_name}': {result.get('error', 'Unknown error')}")
                            sheets_failed.append(sheet_name)
                            
                except Exception as e:
                    logger.warning(f"Failed to populate sheet '{sheet_name}': {e}")
                    sheets_failed.append(sheet_name)
                    qa_report.append(f"❌ '{sheet_name}': Exception - {str(e)[:100]}")
            
            # Log comprehensive QA report
            logger.info("=" * 70)
            logger.info("📋 FINAL QA VALIDATION REPORT - LEARN & IMPROVE")
            logger.info("=" * 70)
            for line in qa_report:
                logger.info(line)
            logger.info("-" * 70)
            logger.info(f"📊 SUMMARY: {len(sheets_populated)} populated, {len(sheets_failed)} failed")
            logger.info(f"   Populated sheets: {sheets_populated}")
            if sheets_failed:
                logger.warning(f"   Failed sheets: {sheets_failed}")
            logger.info("=" * 70)
            
            # Save output
            if not output_filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"pacing_report_{aggregation.value}_{timestamp}.xlsx"
            
            output_path = self.outputs_dir / output_filename
            if job_id:
                self._update_job_status(job_id, "processing", 90, "Saving report file...")
                
            logger.info(f"Saving report to: {output_path}")
            wb.save(str(output_path))
            
            # Capture metadata BEFORE closing/deleting
            sheet_names = wb.sheetnames
            
            # Close workbook and free memory
            wb.close()
            del wb
            gc.collect()
            
            # Calculate summary stats
            total_spend = normalized_data['Spend_USD'].sum()
            total_revenue = normalized_data['Revenue_USD'].sum()
            roas = total_revenue / total_spend if total_spend > 0 else 0
            
            summary = {
                "success": True,
                "output_file": str(output_path),
                "records_processed": len(raw_data),
                "daily_records": len(daily_data),
                "weekly_records": len(weekly_data),
                "date_range": {
                    "start": str(campaign_info['start_date']),
                    "end": str(campaign_info['end_date'])
                },
                "platforms": campaign_info['platforms'].split(' / '),
                "total_spend": float(total_spend),
                "total_revenue": float(total_revenue),
                "overall_roas": float(roas),
                "sheets_populated": sheets_populated,
                "template_sheets": sheet_names
            }
            
            if job_id:
                self._update_job_status(job_id, "completed", 100, "Report generated successfully!", result=summary)
                
            logger.success(f"Report generated: {output_path}")
            return summary
            
        except Exception as e:
            logger.exception(f"Report generation failed: {e}")
            if job_id:
                self._update_job_status(job_id, "failed", 0, str(e))
            return {
                "success": False,
                "error": str(e)
            }
    
    def _update_pivot_analysis_channels(
        self, 
        sheet, 
        daily_data: pd.DataFrame
    ) -> Dict[str, Any]:
        """
        Dynamically update Pivot Analysis channel names based on actual data.
        
        Replaces hardcoded channel names in the "BY CHANNEL" section with actual
        unique channels from the data, while preserving SUMIF formula structure.
        
        Args:
            sheet: Excel worksheet object (Pivot Analysis sheet)
            daily_data: DataFrame with campaign data including 'Platform' column
            
        Returns:
            Dict with update results (channels_updated, row_count, etc.)
        """
        try:
            # Extract unique channels from data, sorted alphabetically
            if 'Platform' not in daily_data.columns:
                logger.warning("No 'Platform' column in data - cannot update Pivot Analysis channels")
                return {'success': False, 'error': 'No Platform column'}
            
            actual_channels = sorted(daily_data['Platform'].unique())
            logger.info(f"🔄 Found {len(actual_channels)} unique channels in data: {actual_channels[:5]}...")
            
            # Find the "BY CHANNEL" section header
            by_channel_row = None
            channel_header_row = None
            
            for row_idx in range(1, min(50, sheet.max_row + 1)):  # Search first 50 rows
                cell_value = sheet.cell(row_idx, 1).value
                if cell_value and isinstance(cell_value, str):
                    if 'BY CHANNEL' in cell_value.upper():
                        by_channel_row = row_idx
                        logger.info(f"Found 'BY CHANNEL' section at row {by_channel_row}")
                    # Look for header row with "Channel" and "Total Spend"
                    elif by_channel_row and 'CHANNEL' in cell_value.upper():
                        next_cell = sheet.cell(row_idx, 2).value
                        if next_cell and 'SPEND' in str(next_cell).upper():
                            channel_header_row = row_idx
                            logger.info(f"Found channel header row at {channel_header_row}")
                            break
            
            if not channel_header_row:
                logger.warning("Could not find 'BY CHANNEL' section in Pivot Analysis")
                return {'success': False, 'error': 'BY CHANNEL section not found'}
            
            # Determine the data rows (start right after header)
            data_start_row = channel_header_row + 1
            
            # Count existing channel rows (until we hit empty or next section)
            existing_channel_count = 0
            for row_idx in range(data_start_row, min(data_start_row + 50, sheet.max_row + 1)):
                cell_value = sheet.cell(row_idx, 1).value
                # Stop if we hit empty row or next section header
                if not cell_value or (isinstance(cell_value, str) and 'BY ' in cell_value.upper()):
                    break
                existing_channel_count += 1
            
            
            logger.info(f"Found {existing_channel_count} existing channel rows, need {len(actual_channels)}")
            
            # Extract template row (first channel row) for copying
            template_row_idx = data_start_row
            
            # CRITICAL: Unmerge any cells in the channel data area to avoid openpyxl bugs
            # where inserted rows inherit merged properties from shifted headers.
            # Row 14 and 25 were originally merged headers in the template.
            # We do this BEFORE inserting/copying so the target area is clean.
            data_end_row = data_start_row + len(actual_channels) + 20 # Extra buffer
            merged_ranges = list(sheet.merged_cells.ranges)
            for m_range in merged_ranges:
                if m_range.min_row <= data_end_row and m_range.max_row >= data_start_row:
                    logger.info(f"Cleanup: Unmerging range {m_range} BEFORE data operations")
                    sheet.unmerge_cells(str(m_range))

            
            # Use insert_rows to shift BY CAMPAIGN and BY YEAR sections down
            if len(actual_channels) > existing_channel_count:
                from copy import copy
                rows_to_add = len(actual_channels) - existing_channel_count
                insert_at = data_start_row + existing_channel_count
                logger.info(f"Adding {rows_to_add} rows by inserting at row {insert_at} and copying template row {template_row_idx}")
                
                # Insert blank rows to shift sections down
                sheet.insert_rows(insert_at, rows_to_add)
                
                # Copy template row to each inserted row
                for offset in range(rows_to_add):
                    target_row = insert_at + offset
                    for col_idx in range(1, 13):  # Columns A through L
                        source_cell = sheet.cell(template_row_idx, col_idx)
                        target_cell = sheet.cell(target_row, col_idx)
                        
                        # Copy cell value (formula or data)
                        if source_cell.value:
                            target_cell.value = None  # Clear first
                            target_cell.value = source_cell.value
                            # Explicitly set data_type for formulas
                            if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                                target_cell.data_type = 'f'
                        
                        # Copy cell formatting
                        if source_cell.has_style:
                            target_cell.font = copy(source_cell.font)
                            target_cell.border = copy(source_cell.border)
                            target_cell.fill = copy(source_cell.fill)
                            target_cell.number_format = copy(source_cell.number_format)
                            target_cell.protection = copy(source_cell.protection)
                            target_cell.alignment = copy(source_cell.alignment)
                
                # Delete or skip debug loop
            
            # If we have too many rows, delete extras
            elif len(actual_channels) < existing_channel_count:
                rows_to_delete = existing_channel_count - len(actual_channels)
                delete_from = data_start_row + len(actual_channels)
                logger.info(f"Deleting {rows_to_delete} rows from row {delete_from}")
                sheet.delete_rows(delete_from, rows_to_delete)
            
            # Now update channel names and adjust formula row references
            channels_updated = 0
            
            # Update all channel rows
            for idx, channel in enumerate(actual_channels):
                row_num = data_start_row + idx
                
                # Update channel name in column A
                sheet.cell(row_num, 1).value = channel
                
                # Update row references in formulas (they were copied from template_row_idx)
                # The copied formulas still reference the template row, so we need to update them
                formulas_updated = 0
                for col_idx in range(2, 13):  # Columns B through L
                    cell = sheet.cell(row_num, col_idx)
                    
                    # RECOVERY: If cell is empty (could happen due to insert_rows bugs), 
                    # fetch formula from template row again
                    if not cell.value:
                        source_ref = sheet.cell(template_row_idx, col_idx)
                        if source_ref.value:
                            cell.value = source_ref.value
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                cell.data_type = 'f'

                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        original_formula = str(cell.value)
                        # Replace template row references with current row references
                        updated_formula = original_formula.replace(
                            f'A{template_row_idx}', f'A{row_num}'
                        ).replace(
                            f'B{template_row_idx}', f'B{row_num}'
                        ).replace(
                            f'C{template_row_idx}', f'C{row_num}'
                        ).replace(
                            f'F{template_row_idx}', f'F{row_num}'
                        ).replace(
                            f'G{template_row_idx}', f'G{row_num}'
                        ).replace(
                            f'H{template_row_idx}', f'H{row_num}'
                        )
                        cell.value = None  # Clear first
                        cell.value = updated_formula
                        cell.data_type = 'f'  # Explicitly mark as formula
                        formulas_updated += 1
                
                channels_updated += 1
            
            logger.success(f"✅ Updated {channels_updated} channels in Pivot Analysis")
            return {
                'success': True,
                'channels_updated': channels_updated,
                'channels': actual_channels
            }
            
        except Exception as e:
            logger.exception(f"Error updating Pivot Analysis channels: {e}")
            return {'success': False, 'error': str(e)}
    
    def _identify_sheet_type(self, sheet_name: str) -> Optional[str]:
        """
        Identify if a sheet should receive daily, weekly, monthly data, or be skipped.
        
        Returns:
            'daily', 'weekly', 'monthly', 'formula' (skip), or None (skip)
        """
        name = sheet_name.lower().strip()
        
        # FORMULA-BASED SHEETS: These use SUMIF/SUMIFS referencing other sheets
        # They should NOT be populated with data - their formulas will auto-calculate
        formula_patterns = ["pivot analysis", "pivot table", "pivot summary", 
                           "multi-granularity", "by channel", "by campaign"]
        
        # Check formula sheets FIRST (most specific)
        if any(p in name for p in formula_patterns):
            logger.debug(f"Sheet '{sheet_name}' identified as FORMULA-based (will be skipped)")
            return 'formula'
        
        # DATA SHEETS: These receive actual data rows
        daily_patterns = ["daily", "day", "campaign data", "raw data", "detail", "all data"]
        weekly_patterns = ["weekly", "week", "pacing tracker"]
        # Budget Pacing Dashboard is MONTHLY aggregation
        monthly_patterns = ["monthly", "month", "budget pacing", "pacing breakdown", "pacing dashboard", 
                          "summary", "overview", "performance"]
        
        if any(p in name for p in daily_patterns):
            logger.debug(f"Sheet '{sheet_name}' identified as DAILY")
            return 'daily'
        if any(p in name for p in weekly_patterns):
            logger.debug(f"Sheet '{sheet_name}' identified as WEEKLY")
            return 'weekly'
        if any(p in name for p in monthly_patterns):
            logger.debug(f"Sheet '{sheet_name}' identified as MONTHLY")
            return 'monthly'
        
        logger.warning(f"⚠️ Sheet '{sheet_name}' not matched to any type - will be skipped")
        return None

    def _detect_data_sheets(self, wb) -> Dict[str, Optional[str]]:
        """
        Legacy method kept for compatibility.
        """
        return {
            "daily": next((n for n in wb.sheetnames if self._identify_sheet_type(n) == 'daily'), wb.sheetnames[0] if wb.sheetnames else None),
            "weekly": next((n for n in wb.sheetnames if self._identify_sheet_type(n) == 'weekly'), None)
        }
