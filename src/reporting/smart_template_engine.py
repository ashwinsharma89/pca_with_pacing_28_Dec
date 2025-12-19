"""
Smart Template Engine - Automatically understands and populates ANY template

Features:
- Detects template structure automatically
- Identifies data fields by context (column headers, labels)
- Supports day-level, week-level, month-level tracking
- Intelligently maps data to template fields
- Handles tables, charts, and summary sections
- Works with any template format

Author: PCA Agent
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Set
from datetime import datetime, timedelta
from dataclasses import dataclass
import re

import pandas as pd
import numpy as np
from loguru import logger
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@dataclass
class TemplateField:
    """Represents a field detected in the template."""
    name: str
    location: str  # Sheet!Cell
    field_type: str  # placeholder, table_header, label, formula
    context: str  # Surrounding text for context
    sheet_name: str
    row: int
    col: int
    original_value: Any = None


@dataclass
class TemplateTable:
    """Represents a data table detected in the template."""
    sheet_name: str
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    headers: List[str]
    table_type: str  # daily, weekly, monthly, summary


class SmartTemplateEngine:
    """
    Intelligent template engine that automatically understands template structure.
    
    Capabilities:
    - Auto-detects placeholders, tables, and data fields
    - Understands context from labels and headers
    - Supports day-level granular tracking
    - Maps data intelligently based on field names
    - Preserves all formatting and formulas
    """
    
    # Common field patterns to detect
    FIELD_PATTERNS = {
        'date': r'date|day|period|week|month|year',
        'spend': r'spend|cost|investment|budget',
        'revenue': r'revenue|sales|income',
        'impressions': r'impression|impr|views',
        'clicks': r'click',
        'conversions': r'conversion|conv|leads',
        'ctr': r'ctr|click.?through|click.?rate',
        'cpc': r'cpc|cost.?per.?click',
        'cpm': r'cpm|cost.?per.?(mille|thousand)',
        'cpa': r'cpa|cost.?per.?(acquisition|action|conversion)',
        'roas': r'roas|return.?on.?ad',
        'campaign': r'campaign|camp',
        'platform': r'platform|channel|source',
        'status': r'status|state|pacing'
    }
    
    def __init__(self):
        """Initialize smart template engine."""
        self.detected_fields: List[TemplateField] = []
        self.detected_tables: List[TemplateTable] = []
        self.template_structure: Dict[str, Any] = {}
        
    def analyze_template(self, template_path: str) -> Dict[str, Any]:
        """
        Automatically analyze template structure.
        
        Args:
            template_path: Path to Excel template
        
        Returns:
            Dictionary with template analysis
        """
        logger.info(f"Analyzing template: {template_path}")
        
        wb = openpyxl.load_workbook(template_path)
        
        analysis = {
            'file_name': Path(template_path).name,
            'sheets': [],
            'fields': [],
            'tables': [],
            'supports_daily': False,
            'supports_weekly': False,
            'supports_monthly': False,
            'field_mapping': {}
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            sheet_analysis = {
                'name': sheet_name,
                'type': self._detect_sheet_type(sheet, sheet_name),
                'fields': [],
                'tables': []
            }
            
            # Scan for fields and tables
            self._scan_sheet(sheet, sheet_name)
            
            # Detect tables
            tables = self._detect_tables(sheet, sheet_name)
            self.detected_tables.extend(tables)
            sheet_analysis['tables'] = [self._table_to_dict(t) for t in tables]
            
            # Check granularity support
            if any('day' in t.table_type.lower() or 'daily' in t.table_type.lower() for t in tables):
                analysis['supports_daily'] = True
            if any('week' in t.table_type.lower() for t in tables):
                analysis['supports_weekly'] = True
            if any('month' in t.table_type.lower() for t in tables):
                analysis['supports_monthly'] = True
            
            analysis['sheets'].append(sheet_analysis)
        
        # Organize detected fields
        analysis['fields'] = [self._field_to_dict(f) for f in self.detected_fields]
        analysis['tables'] = [self._table_to_dict(t) for t in self.detected_tables]
        
        # Create intelligent field mapping
        analysis['field_mapping'] = self._create_field_mapping()
        
        self.template_structure = analysis
        
        logger.info(f"Template analysis complete: {len(self.detected_fields)} fields, {len(self.detected_tables)} tables")
        
        return analysis
    
    def _scan_sheet(self, sheet, sheet_name: str):
        """Scan sheet for fields and patterns."""
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    # Check for placeholders
                    if isinstance(cell.value, str):
                        if any(pattern in str(cell.value) for pattern in ['{{', '{', '[', '<']):
                            field = TemplateField(
                                name=cell.value,
                                location=f"{sheet_name}!{cell.coordinate}",
                                field_type='placeholder',
                                context=self._get_cell_context(sheet, row_idx, col_idx),
                                sheet_name=sheet_name,
                                row=row_idx,
                                col=col_idx,
                                original_value=cell.value
                            )
                            self.detected_fields.append(field)
                        
                        # Check for labeled fields (e.g., "Total Spend:")
                        elif ':' in str(cell.value) or self._is_label(cell.value):
                            # Check if next cell is empty (data field)
                            next_cell = sheet.cell(row=row_idx, column=col_idx + 1)
                            if not next_cell.value or str(next_cell.value).strip() == '':
                                field = TemplateField(
                                    name=cell.value.replace(':', '').strip(),
                                    location=f"{sheet_name}!{next_cell.coordinate}",
                                    field_type='label',
                                    context=cell.value,
                                    sheet_name=sheet_name,
                                    row=row_idx,
                                    col=col_idx + 1,
                                    original_value=next_cell.value
                                )
                                self.detected_fields.append(field)
    
    def _detect_tables(self, sheet, sheet_name: str) -> List[TemplateTable]:
        """Detect data tables in sheet."""
        tables = []
        
        # Look for rows with multiple non-empty cells (potential headers)
        for row_idx in range(1, min(sheet.max_row + 1, 50)):  # Check first 50 rows
            row_values = [sheet.cell(row=row_idx, column=col).value 
                         for col in range(1, min(sheet.max_column + 1, 20))]
            
            # Count non-empty cells
            non_empty = sum(1 for v in row_values if v and str(v).strip())
            
            # If row has 3+ non-empty cells, might be a header
            if non_empty >= 3:
                # Check if values look like headers
                if self._looks_like_headers(row_values):
                    # Found potential table
                    headers = [str(v).strip() for v in row_values if v and str(v).strip()]
                    
                    # Determine table extent
                    start_col = next(i + 1 for i, v in enumerate(row_values) if v and str(v).strip())
                    end_col = start_col + len(headers) - 1
                    
                    # Find end row (where data stops)
                    end_row = self._find_table_end(sheet, row_idx + 1, start_col, end_col)
                    
                    if end_row > row_idx + 1:  # At least one data row
                        table = TemplateTable(
                            sheet_name=sheet_name,
                            start_row=row_idx,
                            start_col=start_col,
                            end_row=end_row,
                            end_col=end_col,
                            headers=headers,
                            table_type=self._detect_table_type(headers, sheet_name)
                        )
                        tables.append(table)
        
        return tables
    
    def _looks_like_headers(self, values: List[Any]) -> bool:
        """Check if values look like table headers."""
        non_empty = [v for v in values if v and str(v).strip()]
        
        if len(non_empty) < 3:
            return False
        
        # Check if values contain common header keywords
        header_keywords = ['date', 'day', 'campaign', 'spend', 'clicks', 'impressions', 
                          'conversions', 'revenue', 'ctr', 'cpc', 'roas', 'platform', 
                          'budget', 'status', 'total', 'week', 'month']
        
        matches = sum(1 for v in non_empty 
                     if any(keyword in str(v).lower() for keyword in header_keywords))
        
        return matches >= 2
    
    def _find_table_end(self, sheet, start_row: int, start_col: int, end_col: int) -> int:
        """Find the last row of a table."""
        for row_idx in range(start_row, min(sheet.max_row + 1, start_row + 1000)):
            # Check if row is empty
            row_values = [sheet.cell(row=row_idx, column=col).value 
                         for col in range(start_col, end_col + 1)]
            
            non_empty = sum(1 for v in row_values if v and str(v).strip())
            
            if non_empty == 0:
                return row_idx - 1
        
        return min(sheet.max_row, start_row + 999)
    
    def _detect_table_type(self, headers: List[str], sheet_name: str) -> str:
        """Detect if table is daily, weekly, monthly, or summary."""
        headers_lower = ' '.join(headers).lower()
        sheet_lower = sheet_name.lower()
        
        if 'day' in headers_lower or 'daily' in headers_lower or 'daily' in sheet_lower:
            return 'daily'
        elif 'week' in headers_lower or 'weekly' in sheet_lower:
            return 'weekly'
        elif 'month' in headers_lower or 'monthly' in sheet_lower:
            return 'monthly'
        else:
            return 'summary'
    
    def _detect_sheet_type(self, sheet, sheet_name: str) -> str:
        """Detect the purpose of a sheet."""
        name_lower = sheet_name.lower()
        
        if any(word in name_lower for word in ['daily', 'day']):
            return 'daily_tracking'
        elif any(word in name_lower for word in ['weekly', 'week']):
            return 'weekly_tracking'
        elif any(word in name_lower for word in ['monthly', 'month']):
            return 'monthly_tracking'
        elif any(word in name_lower for word in ['summary', 'overview', 'dashboard']):
            return 'summary'
        elif any(word in name_lower for word in ['data', 'raw']):
            return 'data'
        else:
            return 'general'
    
    def _get_cell_context(self, sheet, row: int, col: int, radius: int = 2) -> str:
        """Get surrounding context for a cell."""
        context_parts = []
        
        # Check left
        for c in range(max(1, col - radius), col):
            val = sheet.cell(row=row, column=c).value
            if val:
                context_parts.append(str(val))
        
        # Check above
        for r in range(max(1, row - radius), row):
            val = sheet.cell(row=r, column=col).value
            if val:
                context_parts.append(str(val))
        
        return ' '.join(context_parts)
    
    def _is_label(self, value: str) -> bool:
        """Check if value is a label."""
        value_lower = str(value).lower().strip()
        
        # Common label patterns
        label_patterns = [
            'total', 'average', 'sum', 'count', 'overall',
            'campaign', 'date', 'period', 'status'
        ]
        
        return any(pattern in value_lower for pattern in label_patterns)
    
    def _create_field_mapping(self) -> Dict[str, str]:
        """Create intelligent mapping from detected fields to data columns."""
        mapping = {}
        
        for field in self.detected_fields:
            field_name_lower = field.name.lower().replace('{{', '').replace('}}', '').replace(':', '').strip()
            
            # Try to match to standard metrics
            for metric_type, pattern in self.FIELD_PATTERNS.items():
                if re.search(pattern, field_name_lower, re.IGNORECASE):
                    mapping[field.name] = metric_type
                    break
        
        return mapping
    
    def _field_to_dict(self, field: TemplateField) -> Dict:
        """Convert TemplateField to dictionary."""
        return {
            'name': field.name,
            'location': field.location,
            'type': field.field_type,
            'context': field.context
        }
    
    def _table_to_dict(self, table: TemplateTable) -> Dict:
        """Convert TemplateTable to dictionary."""
        return {
            'sheet': table.sheet_name,
            'range': f"{get_column_letter(table.start_col)}{table.start_row}:{get_column_letter(table.end_col)}{table.end_row}",
            'headers': table.headers,
            'type': table.table_type,
            'rows': table.end_row - table.start_row
        }
    
    def populate_template(
        self,
        template_path: str,
        data: pd.DataFrame,
        output_path: str,
        date_column: str = 'Date'
    ) -> str:
        """
        Intelligently populate template with data.
        
        Args:
            template_path: Path to template
            data: Campaign data
            output_path: Where to save populated template
            date_column: Name of date column in data
        
        Returns:
            Path to populated template
        """
        logger.info(f"Populating template with {len(data)} rows of data")
        
        # Analyze template if not already done
        if not self.template_structure:
            self.analyze_template(template_path)
        
        # Load template
        wb = openpyxl.load_workbook(template_path)
        
        # Populate each detected table
        for table in self.detected_tables:
            self._populate_table(wb, table, data, date_column)
        
        # Populate individual fields
        for field in self.detected_fields:
            if field.field_type == 'placeholder':
                self._populate_field(wb, field, data)
        
        # Save populated template
        wb.save(output_path)
        logger.info(f"Template populated and saved to: {output_path}")
        
        return output_path
    
    def _populate_table(
        self,
        wb: openpyxl.Workbook,
        table: TemplateTable,
        data: pd.DataFrame,
        date_column: str
    ):
        """Populate a detected table with data."""
        sheet = wb[table.sheet_name]
        
        logger.info(f"Populating {table.table_type} table in {table.sheet_name}")
        
        # Map table headers to data columns
        column_mapping = self._map_table_columns(table.headers, data.columns)
        
        # Prepare data based on table type
        if table.table_type == 'daily':
            # Group by date
            if date_column in data.columns:
                daily_data = data.groupby(date_column).agg({
                    col: 'sum' if col in ['Spend', 'Clicks', 'Impressions', 'Conversions', 'Revenue'] else 'first'
                    for col in data.columns if col != date_column
                }).reset_index()
            else:
                daily_data = data
        elif table.table_type == 'weekly':
            # Aggregate by week
            daily_data = self._aggregate_weekly(data, date_column)
        else:
            # Summary data
            daily_data = data
        
        # Populate rows
        data_row = table.start_row + 1
        for idx, row_data in daily_data.iterrows():
            if data_row > table.end_row:
                # Need to insert new rows
                sheet.insert_rows(data_row)
            
            for header_idx, header in enumerate(table.headers):
                col = table.start_col + header_idx
                
                # Find matching data column
                data_col = column_mapping.get(header)
                if data_col and data_col in row_data:
                    value = row_data[data_col]
                    
                    # Format value
                    if isinstance(value, (int, float)):
                        value = round(value, 2)
                    
                    sheet.cell(row=data_row, column=col, value=value)
            
            data_row += 1
    
    def _populate_field(self, wb: openpyxl.Workbook, field: TemplateField, data: pd.DataFrame):
        """Populate an individual field."""
        sheet = wb[field.sheet_name]
        
        # Get mapped metric type
        metric_type = self.template_structure['field_mapping'].get(field.name)
        
        if metric_type:
            # Calculate value based on metric type
            value = self._calculate_metric(data, metric_type)
            
            # Update cell
            cell = sheet.cell(row=field.row, column=field.col)
            
            if field.field_type == 'placeholder':
                # Replace placeholder
                if cell.value and isinstance(cell.value, str):
                    cell.value = cell.value.replace(field.name, str(value))
                else:
                    cell.value = value
            else:
                cell.value = value
    
    def _map_table_columns(self, headers: List[str], data_columns: List[str]) -> Dict[str, str]:
        """Map table headers to data columns."""
        mapping = {}
        
        for header in headers:
            header_lower = header.lower().strip()
            
            # Try exact match first
            for col in data_columns:
                if col.lower() == header_lower:
                    mapping[header] = col
                    break
            
            # Try pattern matching
            if header not in mapping:
                for metric_type, pattern in self.FIELD_PATTERNS.items():
                    if re.search(pattern, header_lower, re.IGNORECASE):
                        # Find matching column in data
                        for col in data_columns:
                            if re.search(pattern, col.lower(), re.IGNORECASE):
                                mapping[header] = col
                                break
                        break
        
        return mapping
    
    def _calculate_metric(self, data: pd.DataFrame, metric_type: str) -> Any:
        """Calculate metric value from data."""
        metric_columns = {
            'spend': ['Spend', 'Cost', 'Investment'],
            'revenue': ['Revenue', 'Sales'],
            'clicks': ['Clicks'],
            'impressions': ['Impressions', 'Impr'],
            'conversions': ['Conversions', 'Conv', 'Leads']
        }
        
        # Find matching column
        for col in metric_columns.get(metric_type, []):
            if col in data.columns:
                return data[col].sum()
        
        return 0
    
    def _aggregate_weekly(self, data: pd.DataFrame, date_column: str) -> pd.DataFrame:
        """Aggregate data by week."""
        if date_column not in data.columns:
            return data
        
        data_copy = data.copy()
        data_copy[date_column] = pd.to_datetime(data_copy[date_column])
        data_copy['Week'] = data_copy[date_column].dt.to_period('W')
        
        weekly = data_copy.groupby('Week').agg({
            col: 'sum' if col in ['Spend', 'Clicks', 'Impressions', 'Conversions', 'Revenue'] else 'first'
            for col in data.columns if col not in [date_column, 'Week']
        }).reset_index()
        
        return weekly


def main():
    """Test smart template engine."""
    engine = SmartTemplateEngine()
    
    # Analyze a template
    analysis = engine.analyze_template('templates/test_template.xlsx')
    
    print("Template Analysis:")
    print(f"  Supports Daily: {analysis['supports_daily']}")
    print(f"  Supports Weekly: {analysis['supports_weekly']}")
    print(f"  Fields: {len(analysis['fields'])}")
    print(f"  Tables: {len(analysis['tables'])}")


if __name__ == "__main__":
    main()
