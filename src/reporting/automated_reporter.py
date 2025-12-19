"""
Automated Daily/Weekly Reporting System

Automatically:
- Fetches campaign data
- Calculates KPIs (CTR, ROAS, CPC, CPM)
- Tracks budget utilization
- Generates reports
- Sends notifications

Author: PCA Agent
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime, timedelta
from dataclasses import dataclass, field
import json

import pandas as pd
import numpy as np
from loguru import logger
import schedule
import time

# Add project root to path
current_dir = Path(__file__).parent
project_root = current_dir.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))


@dataclass
class CampaignBudget:
    """Campaign budget configuration."""
    campaign_id: str
    campaign_name: str
    total_budget: float
    start_date: datetime
    end_date: datetime
    daily_budget: Optional[float] = None
    alert_threshold: float = 0.8  # Alert at 80% budget spent
    
    def __post_init__(self):
        """Calculate daily budget if not provided."""
        if self.daily_budget is None:
            days = (self.end_date - self.start_date).days + 1
            self.daily_budget = self.total_budget / days if days > 0 else 0


@dataclass
class CampaignMetrics:
    """Calculated campaign metrics."""
    date: datetime
    campaign_id: str
    campaign_name: str
    
    # Raw metrics
    spend: float = 0.0
    impressions: int = 0
    clicks: int = 0
    conversions: int = 0
    revenue: float = 0.0
    
    # Calculated KPIs
    ctr: float = 0.0  # Click-Through Rate
    cpc: float = 0.0  # Cost Per Click
    cpm: float = 0.0  # Cost Per Mille (1000 impressions)
    cpa: float = 0.0  # Cost Per Acquisition
    roas: float = 0.0  # Return on Ad Spend
    conversion_rate: float = 0.0
    
    # Budget tracking
    budget_total: float = 0.0
    budget_spent: float = 0.0
    budget_remaining: float = 0.0
    budget_spent_pct: float = 0.0
    budget_remaining_pct: float = 0.0
    daily_budget: float = 0.0
    days_remaining: int = 0
    
    def calculate_kpis(self):
        """Calculate all KPI metrics."""
        # CTR = (Clicks / Impressions) * 100
        self.ctr = (self.clicks / self.impressions * 100) if self.impressions > 0 else 0.0
        
        # CPC = Spend / Clicks
        self.cpc = (self.spend / self.clicks) if self.clicks > 0 else 0.0
        
        # CPM = (Spend / Impressions) * 1000
        self.cpm = (self.spend / self.impressions * 1000) if self.impressions > 0 else 0.0
        
        # CPA = Spend / Conversions
        self.cpa = (self.spend / self.conversions) if self.conversions > 0 else 0.0
        
        # ROAS = Revenue / Spend
        self.roas = (self.revenue / self.spend) if self.spend > 0 else 0.0
        
        # Conversion Rate = (Conversions / Clicks) * 100
        self.conversion_rate = (self.conversions / self.clicks * 100) if self.clicks > 0 else 0.0
    
    def calculate_budget_metrics(self, budget: CampaignBudget):
        """Calculate budget utilization metrics."""
        self.budget_total = budget.total_budget
        self.budget_spent = self.spend
        self.budget_remaining = self.budget_total - self.budget_spent
        self.budget_spent_pct = (self.budget_spent / self.budget_total * 100) if self.budget_total > 0 else 0.0
        self.budget_remaining_pct = 100 - self.budget_spent_pct
        self.daily_budget = budget.daily_budget or 0.0
        
        # Calculate days remaining
        today = datetime.now().date()
        end_date = budget.end_date.date() if isinstance(budget.end_date, datetime) else budget.end_date
        self.days_remaining = (end_date - today).days
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            'date': self.date.strftime('%Y-%m-%d'),
            'campaign_id': self.campaign_id,
            'campaign_name': self.campaign_name,
            'spend': round(self.spend, 2),
            'impressions': self.impressions,
            'clicks': self.clicks,
            'conversions': self.conversions,
            'revenue': round(self.revenue, 2),
            'ctr': round(self.ctr, 2),
            'cpc': round(self.cpc, 2),
            'cpm': round(self.cpm, 2),
            'cpa': round(self.cpa, 2),
            'roas': round(self.roas, 2),
            'conversion_rate': round(self.conversion_rate, 2),
            'budget_total': round(self.budget_total, 2),
            'budget_spent': round(self.budget_spent, 2),
            'budget_remaining': round(self.budget_remaining, 2),
            'budget_spent_pct': round(self.budget_spent_pct, 2),
            'budget_remaining_pct': round(self.budget_remaining_pct, 2),
            'daily_budget': round(self.daily_budget, 2),
            'days_remaining': self.days_remaining
        }


class AutomatedReporter:
    """
    Automated reporting system for daily/weekly campaign reports.
    
    Features:
    - Automatic data fetching
    - KPI calculations
    - Budget tracking
    - Alert generation
    - Report generation
    """
    
    def __init__(self, config_path: Optional[str] = None):
        """
        Initialize automated reporter.
        
        Args:
            config_path: Path to configuration file
        """
        self.config_path = config_path or "config/reporting_config.json"
        self.config = self._load_config()
        self.budgets: Dict[str, CampaignBudget] = {}
        self.historical_data: List[CampaignMetrics] = []
        
        # Output directories
        self.reports_dir = Path("reports/automated")
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        
        logger.info("Automated Reporter initialized")
    
    def _load_config(self) -> Dict[str, Any]:
        """Load configuration from file."""
        config_path = Path(self.config_path)
        if config_path.exists():
            with open(config_path, 'r') as f:
                return json.load(f)
        
        # Default configuration
        return {
            'data_source': 'csv',  # csv, api, database
            'data_path': 'data/campaign_data.csv',
            'schedule': {
                'daily_report_time': '09:00',
                'weekly_report_day': 'Monday',
                'weekly_report_time': '10:00'
            },
            'alerts': {
                'budget_threshold': 0.8,
                'performance_threshold': {
                    'min_roas': 2.0,
                    'max_cpa': 50.0,
                    'min_ctr': 1.0
                }
            },
            'notifications': {
                'email_enabled': False,
                'slack_enabled': False
            }
        }
    
    def load_budgets(self, budgets_file: str):
        """
        Load campaign budgets from file.
        
        Args:
            budgets_file: Path to budgets CSV/JSON file
        """
        budgets_path = Path(budgets_file)
        
        if budgets_path.suffix == '.csv':
            df = pd.read_csv(budgets_path)
            for _, row in df.iterrows():
                budget = CampaignBudget(
                    campaign_id=str(row['campaign_id']),
                    campaign_name=row['campaign_name'],
                    total_budget=float(row['total_budget']),
                    start_date=pd.to_datetime(row['start_date']),
                    end_date=pd.to_datetime(row['end_date']),
                    daily_budget=float(row.get('daily_budget', 0)) or None,
                    alert_threshold=float(row.get('alert_threshold', 0.8))
                )
                self.budgets[budget.campaign_id] = budget
        
        elif budgets_path.suffix == '.json':
            with open(budgets_path, 'r') as f:
                budgets_data = json.load(f)
                for budget_dict in budgets_data:
                    budget = CampaignBudget(
                        campaign_id=budget_dict['campaign_id'],
                        campaign_name=budget_dict['campaign_name'],
                        total_budget=budget_dict['total_budget'],
                        start_date=pd.to_datetime(budget_dict['start_date']),
                        end_date=pd.to_datetime(budget_dict['end_date']),
                        daily_budget=budget_dict.get('daily_budget'),
                        alert_threshold=budget_dict.get('alert_threshold', 0.8)
                    )
                    self.budgets[budget.campaign_id] = budget
        
        logger.info(f"Loaded {len(self.budgets)} campaign budgets")
    
    def fetch_campaign_data(self, date: Optional[datetime] = None) -> pd.DataFrame:
        """
        Fetch campaign data for a specific date.
        
        Args:
            date: Date to fetch data for (default: today)
        
        Returns:
            DataFrame with campaign data
        """
        if date is None:
            date = datetime.now()
        
        data_source = self.config.get('data_source', 'csv')
        
        if data_source == 'csv':
            data_path = Path(self.config.get('data_path', 'data/campaign_data.csv'))
            if data_path.exists():
                df = pd.read_csv(data_path)
                df['Date'] = pd.to_datetime(df['Date'])
                
                # Filter for specific date
                df_filtered = df[df['Date'].dt.date == date.date()]
                return df_filtered
            else:
                logger.warning(f"Data file not found: {data_path}")
                return pd.DataFrame()
        
        elif data_source == 'api':
            # TODO: Implement API data fetching
            logger.info("API data fetching not yet implemented")
            return pd.DataFrame()
        
        elif data_source == 'database':
            # TODO: Implement database data fetching
            logger.info("Database data fetching not yet implemented")
            return pd.DataFrame()
        
        return pd.DataFrame()
    
    def process_campaign_data(self, df: pd.DataFrame, date: datetime) -> List[CampaignMetrics]:
        """
        Process raw campaign data and calculate metrics.
        
        Args:
            df: Raw campaign data
            date: Report date
        
        Returns:
            List of CampaignMetrics objects
        """
        metrics_list = []
        
        # Group by campaign
        if 'Campaign_ID' in df.columns:
            grouped = df.groupby('Campaign_ID')
        elif 'Campaign' in df.columns:
            grouped = df.groupby('Campaign')
        else:
            logger.warning("No campaign identifier found in data")
            return metrics_list
        
        for campaign_id, group in grouped:
            # Aggregate metrics
            metrics = CampaignMetrics(
                date=date,
                campaign_id=str(campaign_id),
                campaign_name=group['Campaign_Name'].iloc[0] if 'Campaign_Name' in group.columns else str(campaign_id),
                spend=group['Spend'].sum() if 'Spend' in group.columns else 0.0,
                impressions=int(group['Impressions'].sum()) if 'Impressions' in group.columns else 0,
                clicks=int(group['Clicks'].sum()) if 'Clicks' in group.columns else 0,
                conversions=int(group['Conversions'].sum()) if 'Conversions' in group.columns else 0,
                revenue=group['Revenue'].sum() if 'Revenue' in group.columns else 0.0
            )
            
            # Calculate KPIs
            metrics.calculate_kpis()
            
            # Calculate budget metrics if budget exists
            if campaign_id in self.budgets:
                metrics.calculate_budget_metrics(self.budgets[campaign_id])
            
            metrics_list.append(metrics)
        
        return metrics_list
    
    def generate_daily_report(self, date: Optional[datetime] = None) -> Dict[str, Any]:
        """
        Generate daily campaign report.
        
        Args:
            date: Report date (default: today)
        
        Returns:
            Report dictionary
        """
        if date is None:
            date = datetime.now()
        
        logger.info(f"Generating daily report for {date.strftime('%Y-%m-%d')}")
        
        # Fetch data
        df = self.fetch_campaign_data(date)
        
        if df.empty:
            logger.warning("No data available for daily report")
            return {'status': 'no_data', 'date': date.strftime('%Y-%m-%d')}
        
        # Process data
        metrics_list = self.process_campaign_data(df, date)
        
        # Generate report
        report = {
            'report_type': 'daily',
            'date': date.strftime('%Y-%m-%d'),
            'generated_at': datetime.now().isoformat(),
            'campaigns': [m.to_dict() for m in metrics_list],
            'summary': self._calculate_summary(metrics_list),
            'alerts': self._generate_alerts(metrics_list)
        }
        
        # Save report
        self._save_report(report, 'daily', date)
        
        # Store in historical data
        self.historical_data.extend(metrics_list)
        
        logger.info(f"Daily report generated with {len(metrics_list)} campaigns")
        
        return report
    
    def generate_weekly_report(self, end_date: Optional[datetime] = None) -> Dict[str, Any]:
        """
        Generate weekly campaign report (last 7 days).
        
        Args:
            end_date: End date for report (default: today)
        
        Returns:
            Report dictionary
        """
        if end_date is None:
            end_date = datetime.now()
        
        start_date = end_date - timedelta(days=6)
        
        logger.info(f"Generating weekly report for {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        
        # Fetch data for each day
        all_metrics = []
        for i in range(7):
            current_date = start_date + timedelta(days=i)
            df = self.fetch_campaign_data(current_date)
            if not df.empty:
                metrics = self.process_campaign_data(df, current_date)
                all_metrics.extend(metrics)
        
        if not all_metrics:
            logger.warning("No data available for weekly report")
            return {'status': 'no_data', 'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"}
        
        # Aggregate weekly metrics by campaign
        weekly_metrics = self._aggregate_weekly_metrics(all_metrics)
        
        # Generate report
        report = {
            'report_type': 'weekly',
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'generated_at': datetime.now().isoformat(),
            'campaigns': [m.to_dict() for m in weekly_metrics],
            'summary': self._calculate_summary(weekly_metrics),
            'trends': self._calculate_trends(all_metrics),
            'alerts': self._generate_alerts(weekly_metrics)
        }
        
        # Save report
        self._save_report(report, 'weekly', end_date)
        
        logger.info(f"Weekly report generated with {len(weekly_metrics)} campaigns")
        
        return report
    
    def _aggregate_weekly_metrics(self, metrics_list: List[CampaignMetrics]) -> List[CampaignMetrics]:
        """Aggregate daily metrics into weekly metrics."""
        # Group by campaign
        campaign_groups = {}
        for metric in metrics_list:
            if metric.campaign_id not in campaign_groups:
                campaign_groups[metric.campaign_id] = []
            campaign_groups[metric.campaign_id].append(metric)
        
        # Aggregate each campaign
        weekly_metrics = []
        for campaign_id, daily_metrics in campaign_groups.items():
            # Sum raw metrics
            weekly = CampaignMetrics(
                date=daily_metrics[-1].date,  # Use last date
                campaign_id=campaign_id,
                campaign_name=daily_metrics[0].campaign_name,
                spend=sum(m.spend for m in daily_metrics),
                impressions=sum(m.impressions for m in daily_metrics),
                clicks=sum(m.clicks for m in daily_metrics),
                conversions=sum(m.conversions for m in daily_metrics),
                revenue=sum(m.revenue for m in daily_metrics)
            )
            
            # Recalculate KPIs
            weekly.calculate_kpis()
            
            # Calculate budget metrics
            if campaign_id in self.budgets:
                weekly.calculate_budget_metrics(self.budgets[campaign_id])
            
            weekly_metrics.append(weekly)
        
        return weekly_metrics
    
    def _calculate_summary(self, metrics_list: List[CampaignMetrics]) -> Dict[str, Any]:
        """Calculate summary statistics across all campaigns."""
        if not metrics_list:
            return {}
        
        total_spend = sum(m.spend for m in metrics_list)
        total_revenue = sum(m.revenue for m in metrics_list)
        total_impressions = sum(m.impressions for m in metrics_list)
        total_clicks = sum(m.clicks for m in metrics_list)
        total_conversions = sum(m.conversions for m in metrics_list)
        
        return {
            'total_campaigns': len(metrics_list),
            'total_spend': round(total_spend, 2),
            'total_revenue': round(total_revenue, 2),
            'total_impressions': total_impressions,
            'total_clicks': total_clicks,
            'total_conversions': total_conversions,
            'avg_ctr': round((total_clicks / total_impressions * 100) if total_impressions > 0 else 0.0, 2),
            'avg_cpc': round((total_spend / total_clicks) if total_clicks > 0 else 0.0, 2),
            'avg_cpm': round((total_spend / total_impressions * 1000) if total_impressions > 0 else 0.0, 2),
            'avg_cpa': round((total_spend / total_conversions) if total_conversions > 0 else 0.0, 2),
            'overall_roas': round((total_revenue / total_spend) if total_spend > 0 else 0.0, 2),
            'conversion_rate': round((total_conversions / total_clicks * 100) if total_clicks > 0 else 0.0, 2)
        }
    
    def _calculate_trends(self, metrics_list: List[CampaignMetrics]) -> Dict[str, Any]:
        """Calculate week-over-week trends."""
        if len(metrics_list) < 2:
            return {}
        
        # Sort by date
        sorted_metrics = sorted(metrics_list, key=lambda x: x.date)
        
        # Split into first half and second half
        mid_point = len(sorted_metrics) // 2
        first_half = sorted_metrics[:mid_point]
        second_half = sorted_metrics[mid_point:]
        
        # Calculate metrics for each half
        first_spend = sum(m.spend for m in first_half)
        second_spend = sum(m.spend for m in second_half)
        
        first_conversions = sum(m.conversions for m in first_half)
        second_conversions = sum(m.conversions for m in second_half)
        
        first_roas = sum(m.revenue for m in first_half) / first_spend if first_spend > 0 else 0
        second_roas = sum(m.revenue for m in second_half) / second_spend if second_spend > 0 else 0
        
        return {
            'spend_change_pct': round(((second_spend - first_spend) / first_spend * 100) if first_spend > 0 else 0.0, 2),
            'conversions_change_pct': round(((second_conversions - first_conversions) / first_conversions * 100) if first_conversions > 0 else 0.0, 2),
            'roas_change_pct': round(((second_roas - first_roas) / first_roas * 100) if first_roas > 0 else 0.0, 2)
        }
    
    def _generate_alerts(self, metrics_list: List[CampaignMetrics]) -> List[Dict[str, Any]]:
        """Generate alerts based on thresholds."""
        alerts = []
        
        alert_config = self.config.get('alerts', {})
        budget_threshold = alert_config.get('budget_threshold', 0.8)
        perf_thresholds = alert_config.get('performance_threshold', {})
        
        for metric in metrics_list:
            # Budget alerts
            if metric.budget_spent_pct >= budget_threshold * 100:
                alerts.append({
                    'type': 'budget',
                    'severity': 'high' if metric.budget_spent_pct >= 90 else 'medium',
                    'campaign_id': metric.campaign_id,
                    'campaign_name': metric.campaign_name,
                    'message': f"Budget {metric.budget_spent_pct:.1f}% spent (${metric.budget_spent:,.2f} of ${metric.budget_total:,.2f})",
                    'value': metric.budget_spent_pct
                })
            
            # Performance alerts
            min_roas = perf_thresholds.get('min_roas', 2.0)
            if metric.roas > 0 and metric.roas < min_roas:
                alerts.append({
                    'type': 'performance',
                    'severity': 'medium',
                    'campaign_id': metric.campaign_id,
                    'campaign_name': metric.campaign_name,
                    'message': f"ROAS below target: {metric.roas:.2f}x (target: {min_roas}x)",
                    'value': metric.roas
                })
            
            max_cpa = perf_thresholds.get('max_cpa', 50.0)
            if metric.cpa > 0 and metric.cpa > max_cpa:
                alerts.append({
                    'type': 'performance',
                    'severity': 'medium',
                    'campaign_id': metric.campaign_id,
                    'campaign_name': metric.campaign_name,
                    'message': f"CPA above target: ${metric.cpa:.2f} (target: ${max_cpa:.2f})",
                    'value': metric.cpa
                })
            
            min_ctr = perf_thresholds.get('min_ctr', 1.0)
            if metric.ctr > 0 and metric.ctr < min_ctr:
                alerts.append({
                    'type': 'performance',
                    'severity': 'low',
                    'campaign_id': metric.campaign_id,
                    'campaign_name': metric.campaign_name,
                    'message': f"CTR below target: {metric.ctr:.2f}% (target: {min_ctr}%)",
                    'value': metric.ctr
                })
        
        return alerts
    
    def _save_report(self, report: Dict[str, Any], report_type: str, date: datetime):
        """Save report to file."""
        filename = f"{report_type}_report_{date.strftime('%Y%m%d')}.json"
        filepath = self.reports_dir / filename
        
        with open(filepath, 'w') as f:
            json.dump(report, f, indent=2)
        
        logger.info(f"Report saved to {filepath}")
        
        # Also save as CSV for easy viewing
        if report.get('campaigns'):
            df = pd.DataFrame(report['campaigns'])
            csv_filename = f"{report_type}_report_{date.strftime('%Y%m%d')}.csv"
            csv_filepath = self.reports_dir / csv_filename
            df.to_csv(csv_filepath, index=False)
            logger.info(f"Report CSV saved to {csv_filepath}")
    
    def export_to_excel(self, report: Dict[str, Any], output_path: str):
        """
        Export report to formatted Excel file.
        
        Args:
            report: Report dictionary
            output_path: Output file path
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add title
        ws_summary['A1'] = f"{report['report_type'].title()} Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        # Add date
        if report['report_type'] == 'daily':
            ws_summary['A2'] = f"Date: {report['date']}"
        else:
            ws_summary['A2'] = f"Period: {report['start_date']} to {report['end_date']}"
        
        # Add summary metrics
        row = 4
        summary = report.get('summary', {})
        for key, value in summary.items():
            ws_summary[f'A{row}'] = key.replace('_', ' ').title()
            ws_summary[f'B{row}'] = value
            row += 1
        
        # Campaigns sheet
        if report.get('campaigns'):
            ws_campaigns = wb.create_sheet("Campaigns")
            df = pd.DataFrame(report['campaigns'])
            
            # Add headers
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws_campaigns.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Add data
            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws_campaigns.cell(row=row_idx, column=col_idx, value=value)
        
        # Alerts sheet
        if report.get('alerts'):
            ws_alerts = wb.create_sheet("Alerts")
            df_alerts = pd.DataFrame(report['alerts'])
            
            for col_idx, col_name in enumerate(df_alerts.columns, 1):
                cell = ws_alerts.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            for row_idx, row_data in enumerate(df_alerts.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws_alerts.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(output_path)
        logger.info(f"Excel report saved to {output_path}")
    
    def schedule_reports(self):
        """Schedule automated report generation."""
        schedule_config = self.config.get('schedule', {})
        
        # Daily report
        daily_time = schedule_config.get('daily_report_time', '09:00')
        schedule.every().day.at(daily_time).do(self.generate_daily_report)
        logger.info(f"Scheduled daily report at {daily_time}")
        
        # Weekly report
        weekly_day = schedule_config.get('weekly_report_day', 'Monday')
        weekly_time = schedule_config.get('weekly_report_time', '10:00')
        getattr(schedule.every(), weekly_day.lower()).at(weekly_time).do(self.generate_weekly_report)
        logger.info(f"Scheduled weekly report on {weekly_day} at {weekly_time}")
    
    def run_scheduler(self):
        """Run the scheduler loop."""
        logger.info("Starting automated reporting scheduler...")
        
        while True:
            schedule.run_pending()
            time.sleep(60)  # Check every minute


def main():
    """Main entry point for automated reporting."""
    # Initialize reporter
    reporter = AutomatedReporter()
    
    # Load budgets
    budgets_file = "config/campaign_budgets.csv"
    if Path(budgets_file).exists():
        reporter.load_budgets(budgets_file)
    
    # Generate reports on demand
    print("Generating daily report...")
    daily_report = reporter.generate_daily_report()
    
    print("\nGenerating weekly report...")
    weekly_report = reporter.generate_weekly_report()
    
    # Export to Excel
    reporter.export_to_excel(daily_report, "reports/automated/daily_report.xlsx")
    reporter.export_to_excel(weekly_report, "reports/automated/weekly_report.xlsx")
    
    print("\nReports generated successfully!")
    print(f"- Daily report: {len(daily_report.get('campaigns', []))} campaigns")
    print(f"- Weekly report: {len(weekly_report.get('campaigns', []))} campaigns")
    print(f"- Alerts: {len(daily_report.get('alerts', []))} alerts")
    
    # Optionally start scheduler
    # reporter.schedule_reports()
    # reporter.run_scheduler()


if __name__ == "__main__":
    main()
