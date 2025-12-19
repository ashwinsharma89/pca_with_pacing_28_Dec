"""
Test Automated Reporting with Local Template

This script:
1. Generates automated report data (KPIs, budget tracking)
2. Uses your local template
3. Populates template with calculated metrics
4. Creates final report

Usage:
    python test_reporting_with_template.py --template "path/to/your/template.xlsx"
"""

import sys
import argparse
from pathlib import Path
from datetime import datetime
import pandas as pd
import json

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.reporting.automated_reporter import AutomatedReporter, CampaignMetrics, CampaignBudget


def generate_sample_data() -> pd.DataFrame:
    """Generate sample campaign data for testing."""
    data = {
        'Date': ['2024-12-04'] * 5,
        'Campaign_ID': ['CAMP001', 'CAMP002', 'CAMP003', 'CAMP004', 'CAMP005'],
        'Campaign_Name': [
            'Q4 Brand Campaign',
            'Holiday Sales Push',
            'Product Launch',
            'Retargeting Campaign',
            'Lead Generation'
        ],
        'Spend': [1500.50, 2800.75, 950.25, 450.00, 1200.00],
        'Impressions': [45000, 68000, 28000, 15000, 35000],
        'Clicks': [450, 820, 280, 180, 420],
        'Conversions': [25, 45, 18, 12, 28],
        'Revenue': [7500.00, 14000.00, 5400.00, 3600.00, 8400.00]
    }
    return pd.DataFrame(data)


def create_template_mapping(metrics_list, template_path: str) -> dict:
    """
    Create mapping between calculated metrics and template placeholders.
    
    This analyzes your template and suggests mappings.
    """
    import openpyxl
    
    # Load template to detect placeholders
    wb = openpyxl.load_workbook(template_path)
    placeholders = {}
    
    print("\n[ANALYZING] Your template...")
    print(f"   Template: {Path(template_path).name}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")
    
    # Scan for placeholders
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check for placeholder patterns
                    if any(pattern in str(cell.value) for pattern in ['{{', '{', '[', '<']):
                        location = f"{sheet_name}!{cell.coordinate}"
                        placeholders[cell.value] = location
                        print(f"   Found: {cell.value} at {location}")
    
    if not placeholders:
        print("   ⚠ No placeholders found. Will create data mapping...")
    
    # Calculate aggregate metrics
    total_spend = sum(m.spend for m in metrics_list)
    total_revenue = sum(m.revenue for m in metrics_list)
    total_impressions = sum(m.impressions for m in metrics_list)
    total_clicks = sum(m.clicks for m in metrics_list)
    total_conversions = sum(m.conversions for m in metrics_list)
    
    # Create mapping for common placeholders
    mapping = {
        # Totals
        '{{Total_Spend}}': total_spend,
        '{{Total_Revenue}}': total_revenue,
        '{{Total_Impressions}}': total_impressions,
        '{{Total_Clicks}}': total_clicks,
        '{{Total_Conversions}}': total_conversions,
        
        # KPIs
        '{{Overall_ROAS}}': total_revenue / total_spend if total_spend > 0 else 0,
        '{{Avg_CPC}}': total_spend / total_clicks if total_clicks > 0 else 0,
        '{{Avg_CPM}}': (total_spend / total_impressions * 1000) if total_impressions > 0 else 0,
        '{{Avg_CPA}}': total_spend / total_conversions if total_conversions > 0 else 0,
        '{{Avg_CTR}}': (total_clicks / total_impressions * 100) if total_impressions > 0 else 0,
        '{{Conversion_Rate}}': (total_conversions / total_clicks * 100) if total_clicks > 0 else 0,
        
        # Budget (if available)
        '{{Budget_Total}}': sum(m.budget_total for m in metrics_list if m.budget_total > 0),
        '{{Budget_Spent}}': sum(m.budget_spent for m in metrics_list if m.budget_spent > 0),
        '{{Budget_Remaining}}': sum(m.budget_remaining for m in metrics_list if m.budget_remaining > 0),
        '{{Budget_Spent_Pct}}': sum(m.budget_spent_pct for m in metrics_list) / len(metrics_list) if metrics_list else 0,
        
        # Date
        '{{Report_Date}}': datetime.now().strftime('%Y-%m-%d'),
        '{{Report_Month}}': datetime.now().strftime('%B %Y'),
        
        # Campaign counts
        '{{Campaign_Count}}': len(metrics_list),
    }
    
    # Add per-campaign metrics
    for i, metric in enumerate(metrics_list, 1):
        campaign_prefix = f"{{{{Campaign{i}_"
        mapping.update({
            f"{campaign_prefix}Name}}}}": metric.campaign_name,
            f"{campaign_prefix}Spend}}}}": metric.spend,
            f"{campaign_prefix}Clicks}}}}": metric.clicks,
            f"{campaign_prefix}Conversions}}}}": metric.conversions,
            f"{campaign_prefix}ROAS}}}}": metric.roas,
            f"{campaign_prefix}CPC}}}}": metric.cpc,
            f"{campaign_prefix}CTR}}}}": metric.ctr,
        })
    
    return mapping, placeholders


def populate_template(template_path: str, mapping: dict, output_path: str, metrics_df: pd.DataFrame):
    """
    Populate template with calculated metrics.
    
    Args:
        template_path: Path to your template file
        mapping: Dictionary of placeholder -> value mappings
        output_path: Where to save populated report
        metrics_df: DataFrame with all campaign metrics
    """
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment
    
    print("\n[POPULATING] Template...")
    
    # Load template
    wb = openpyxl.load_workbook(template_path)
    
    replacements_made = 0
    
    # Replace placeholders in all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_value = cell.value
                    updated_value = original_value
                    
                    # Replace all matching placeholders
                    for placeholder, value in mapping.items():
                        if placeholder in updated_value:
                            # Format value based on type
                            if isinstance(value, float):
                                if 'ROAS' in placeholder or 'CTR' in placeholder or 'Rate' in placeholder:
                                    formatted_value = f"{value:.2f}"
                                elif 'Pct' in placeholder or 'Percent' in placeholder:
                                    formatted_value = f"{value:.1f}%"
                                else:
                                    formatted_value = f"${value:,.2f}"
                            elif isinstance(value, int):
                                formatted_value = f"{value:,}"
                            else:
                                formatted_value = str(value)
                            
                            updated_value = updated_value.replace(placeholder, formatted_value)
                            replacements_made += 1
                    
                    # Update cell if changed
                    if updated_value != original_value:
                        cell.value = updated_value
    
    print(f"   [OK] Made {replacements_made} replacements")
    
    # Add data sheet with all metrics
    if 'Campaign_Data' not in wb.sheetnames:
        data_sheet = wb.create_sheet('Campaign_Data')
    else:
        data_sheet = wb['Campaign_Data']
        # Clear existing data
        data_sheet.delete_rows(1, data_sheet.max_row)
    
    print("   [OK] Adding Campaign_Data sheet...")
    
    # Add headers with formatting
    for c_idx, col_name in enumerate(metrics_df.columns, 1):
        cell = data_sheet.cell(row=1, column=c_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for r_idx, row in enumerate(dataframe_to_rows(metrics_df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = data_sheet.cell(row=r_idx, column=c_idx, value=value)
            # Format numbers
            if isinstance(value, float):
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for column in data_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        data_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save populated template
    wb.save(output_path)
    print(f"   [OK] Saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Test automated reporting with your template')
    parser.add_argument(
        '--template',
        type=str,
        required=True,
        help='Path to your report template (Excel file)'
    )
    parser.add_argument(
        '--data',
        type=str,
        help='Path to campaign data CSV (optional, will use sample data if not provided)'
    )
    parser.add_argument(
        '--budgets',
        type=str,
        default='config/campaign_budgets.csv',
        help='Path to campaign budgets file'
    )
    parser.add_argument(
        '--output',
        type=str,
        help='Output path for populated report (default: auto-generated)'
    )
    
    args = parser.parse_args()
    
    # Validate template exists
    template_path = Path(args.template)
    if not template_path.exists():
        print(f"❌ Template not found: {args.template}")
        print("\nPlease provide a valid path to your Excel template.")
        print("Example: python test_reporting_with_template.py --template 'C:/Users/yourname/template.xlsx'")
        return
    
    print("=" * 70)
    print("PCA Agent - Automated Reporting with Your Template")
    print("=" * 70)
    
    # Initialize reporter
    print("\n[1] Initializing automated reporter...")
    reporter = AutomatedReporter()
    
    # Load budgets if available
    if Path(args.budgets).exists():
        print(f"   Loading budgets from {args.budgets}...")
        reporter.load_budgets(args.budgets)
        print(f"   [OK] Loaded {len(reporter.budgets)} campaign budgets")
    
    # Load or generate data
    print("\n[2] Loading campaign data...")
    if args.data and Path(args.data).exists():
        print(f"   Loading from {args.data}...")
        df = pd.read_csv(args.data)
    else:
        print("   Using sample data for testing...")
        df = generate_sample_data()
        # Save sample data for reference
        sample_path = 'data/sample_campaign_data.csv'
        Path(sample_path).parent.mkdir(exist_ok=True)
        df.to_csv(sample_path, index=False)
        print(f"   [OK] Sample data saved to: {sample_path}")
    
    print(f"   [OK] Loaded {len(df)} rows")
    
    # Process data and calculate metrics
    print("\n[3] Calculating KPIs...")
    date = datetime.now()
    metrics_list = reporter.process_campaign_data(df, date)
    
    print(f"   [OK] Processed {len(metrics_list)} campaigns")
    print("\n   Campaign Metrics:")
    for metric in metrics_list:
        print(f"   - {metric.campaign_name}")
        print(f"     Spend: ${metric.spend:,.2f} | ROAS: {metric.roas:.2f}x | CTR: {metric.ctr:.2f}%")
        print(f"     CPC: ${metric.cpc:.2f} | CPA: ${metric.cpa:.2f} | Conv Rate: {metric.conversion_rate:.2f}%")
        if metric.budget_total > 0:
            print(f"     Budget: ${metric.budget_spent:,.2f} / ${metric.budget_total:,.2f} ({metric.budget_spent_pct:.1f}%)")
    
    # Create metrics DataFrame for data sheet
    metrics_df = pd.DataFrame([m.to_dict() for m in metrics_list])
    
    # Create mapping
    print("\n[4] Creating template mapping...")
    mapping, detected_placeholders = create_template_mapping(metrics_list, str(template_path))
    
    if detected_placeholders:
        print(f"\n   Detected {len(detected_placeholders)} placeholders in your template")
        print("   Will populate these with calculated metrics")
    else:
        print("\n   No placeholders detected - will add Campaign_Data sheet")
    
    # Generate output path
    if args.output:
        output_path = args.output
    else:
        template_name = template_path.stem
        output_path = f"reports/automated/{template_name}_populated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Populate template
    print("\n[5] Generating report...")
    populate_template(str(template_path), mapping, output_path, metrics_df)
    
    # Summary
    print("\n" + "=" * 70)
    print("[SUCCESS] Report Generated Successfully!")
    print("=" * 70)
    print(f"\nSummary:")
    print(f"   Template Used: {template_path.name}")
    print(f"   Campaigns Processed: {len(metrics_list)}")
    print(f"   Total Spend: ${sum(m.spend for m in metrics_list):,.2f}")
    print(f"   Total Revenue: ${sum(m.revenue for m in metrics_list):,.2f}")
    print(f"   Overall ROAS: {(sum(m.revenue for m in metrics_list) / sum(m.spend for m in metrics_list)):.2f}x")
    print(f"\nOutput: {output_path}")
    print("\nNext Steps:")
    print("   1. Open the generated report in Excel")
    print("   2. Review the populated metrics")
    print("   3. Check the Campaign_Data sheet for detailed data")
    print("   4. Customize your template with more placeholders if needed")
    print("\n" + "=" * 70)


if __name__ == "__main__":
    main()
